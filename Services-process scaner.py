#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NTSE.py — إصدار مُحسَّن وفق طلباتك
التغييرات الرئيسية:
1) عمود signature يعرض بالإنجليزية: "Valid" أو "Invalid".
2) حل مشكلة الكروت الخاصة بالخدمات النشطة والعمليات الجارية والعمليات غير المُوقَّعة بحيث تعرض النتائج عند النقر؛ حيث إذا لم تتوافق البيانات مع الفلاتر يتم عرض رسالة تنبيهية.
3) تصحيح فلترة كرت "العمليات غير مُوقَّعة" بحيث تُظهر العمليات التي لا تحتوي على توقيع صحيح.
4) تعديل دالة تصدير PDF بحيث يتم تصدير البيانات النصية مع الرسوم البيانية (المخططات) المُنشأة بواسطة pyqtgraph.
5) تخزين الكلمات المفتاحية المدخلة تلقائيًا.
6) حذف زر تحميل القوائم.
7) حذف زر "حفظ قائمة الكلمات" واستبداله بزر حذف كل القائمة في نفس الموقع.
8) تعديل تنسيقات الواجهة بحيث تشمل QMessageBox وQComboBox وQAbstractItemView تتأقلم مع الثيم.
9) تعديل فلترة الحساب بحيث إذا اختار المستخدم حسابًا محددًا (مثل User أو SYSTEM) تظهر العمليات والخدمات التابعة لذلك الحساب، وإذا اختار All تظهر كل النتائج.
10) جعل الأداة تعمل على نظامي ويندوز ولينوكس.
11) إضافة زر يحتوي على قائمة منسدلة (أو شريط أدوات عائم) يحتوي على خيارين: "حفظ كـ PDF" و"مقارنة النتائج" مع إضافة دوال لحفظ النتائج في ملف JSON وتحميل النتائج السابقة للمقارنة.
"""

import sys, os, json, hashlib, traceback
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

# ===== اعتمادات =====
try:
    import psutil
except Exception as e:
    raise ImportError("psutil مطلوب: pip install psutil") from e

try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFrame,
        QPushButton, QLabel, QLineEdit, QListWidget, QListWidgetItem, QInputDialog,
        QComboBox, QSpinBox, QCheckBox, QTableView, QHeaderView,
        QAbstractItemView, QMessageBox, QFileDialog, QStatusBar, QProgressBar,
        QSizePolicy, QToolButton, QMenu
    )
    from PyQt5.QtCore import Qt, QThread, pyqtSignal, QEvent, QSortFilterProxyModel, QModelIndex
    from PyQt5.QtGui import QStandardItemModel, QStandardItem, QKeySequence, QPixmap
except Exception as e:
    raise ImportError("PyQt5 مطلوب: pip install PyQt5") from e

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

try:
    import pefile
except Exception:
    pefile = None

try:
    import pyqtgraph as pg
except Exception:
    pg = None

# استيراد مكتبة ReportLab لحفظ النتائج كـ PDF
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except Exception as e:
    raise ImportError("ReportLab مطلوب: pip install reportlab") from e

APP_DIR = Path.home() / ".ntse_ui"
APP_DIR.mkdir(exist_ok=True)
LISTS_FILE = APP_DIR / "lists.json"

# ===== الترجمات =====
L_AR = {
    "title": "أداة فحص النظام — واجهة مُحسّنة",
    "scan": "بدء الفحص",
    "stop": "إيقاف",
    "clear": "مسح",
    "export": "تصدير XLSX",
    "exit": "Exit",
    "theme": "الثيم",
    "lang": "اللغة",
    "criteria": "المعايير",
    "scope": "نطاق الفحص",
    "scope_both": "الخدمات والعمليات",
    "scope_services": "الخدمات فقط",
    "scope_processes": "العمليات فقط",
    "keywords": "كلمات مفتاحية (مطابقة جزئية)",
    "add": "+",
    "remove": "−",
    "delete_all": "حذف كل القائمة",
    "start_type": "نوع بدء الخدمة",
    "acc": "الحساب (لعرض فقط)",
    "days": "عمر الملف (أيام)",
    "use_age": "اعتبار العمر",
    "use_sig": "فحص التوقيع",
    "stats": "إحصاءات الفحص",
    "reasons_chart": "أسباب الاشتباه",
    "results": "نتائج الفحص",
    "filter_hint": "الفلترة (Ctrl+F للإظهار/الإخفاء)",
    "table_headers": [
        "النوع", "الاسم", "الاسم المعروض", "الحالة", "نوع البدء", "الحساب",
        "المسار التنفيذي", "Signature", "أسباب", "تاريخ الإنشاء"
    ],
    "type_service": "خدمة",
    "type_process": "عملية",
    "start_all": "الكل",
    "start_auto": "تلقائي",
    "start_manual": "يدوي",
    "start_disabled": "معطل",
    "running": "جارية",
    "saved": "تم الحفظ",
    "loaded": "تم التحميل",
    "need_openpyxl": "ثبّت openpyxl لتفعيل التصدير: pip install openpyxl",
    "progress": "جاري الفحص... {}/{}",
    "done": "انتهى الفحص: {} عنصر من {}.",
    "search_bar_placeholder": "اكتب الشرط...",
    "col": "العمود",
    "apply": "تطبيق",
    "cards": {
        "procs_running": "العمليات الجارية",
        "svcs_active": "الخدمات النشطة",
        "procs_kw": "عمليات مطابقة للكلمات",
        "svcs_kw": "خدمات مطابقة للكلمات",
        "procs_unsigned": "العمليات غير مُوقَّعة"
    },
    "chips": {
        "procs_running": "العمليات الجارية",
        "svcs_active": "الخدمات النشطة"
    },
    "options": "خيارات",
    "save_pdf": "حفظ كـ PDF",
    "compare": "مقارنة النتائج"
}
L_EN = {
    "title": "System Scanner — Enhanced UI",
    "scan": "Scan",
    "stop": "Stop",
    "clear": "Clear",
    "export": "Export XLSX",
    "exit": "Exit",
    "theme": "Theme",
    "lang": "Language",
    "criteria": "Criteria",
    "scope": "Scope",
    "scope_both": "Services & Processes",
    "scope_services": "Services only",
    "scope_processes": "Processes only",
    "keywords": "Keywords (partial match)",
    "add": "+",
    "remove": "−",
    "delete_all": "Delete All",
    "start_type": "Service start type",
    "acc": "Account (display only)",
    "days": "File age (days)",
    "use_age": "Check age",
    "use_sig": "Check signature",
    "stats": "Scan Stats",
    "reasons_chart": "Suspicion reasons",
    "results": "Scan Results",
    "filter_hint": "Filter (Ctrl+F to toggle)",
    "table_headers": [
        "Type", "Name", "Display name", "State", "Start type", "Account",
        "Executable path", "Signature", "Reasons", "Timestamp"
    ],
    "type_service": "Service",
    "type_process": "Process",
    "start_all": "All",
    "start_auto": "Automatic",
    "start_manual": "Manual",
    "start_disabled": "Disabled",
    "running": "Running",
    "saved": "Saved",
    "loaded": "Loaded",
    "need_openpyxl": "Install openpyxl to enable export: pip install openpyxl",
    "progress": "Scanning... {}/{}",
    "done": "Done: {} of {} items.",
    "search_bar_placeholder": "Enter condition...",
    "col": "Column",
    "apply": "Apply",
    "cards": {
        "procs_running": "Running Processes",
        "svcs_active": "Active Services",
        "procs_kw": "Keyword-matched Processes",
        "svcs_kw": "Keyword-matched Services",
        "procs_unsigned": "Unsigned Processes"
    },
    "chips": {
        "procs_running": "Running Processes",
        "svcs_active": "Active Services"
    },
    "options": "Options",
    "save_pdf": "Save as PDF",
    "compare": "Compare Results"
}

LANG = "ar"
def tr(key: str):
    return (L_AR if LANG=="ar" else L_EN).get(key, key)
def tr_sub(section: str, key: str) -> str:
    d = (L_AR if LANG=="ar" else L_EN).get(section, {})
    return d.get(key, key)
def table_headers() -> List[str]:
    return L_AR["table_headers"] if LANG=="ar" else L_EN["table_headers"]

# ===== أدوات =====
def split_tokens(raw: List[str]) -> List[str]:
    out = []
    for s in raw:
        for part in s.split(','):
            t = part.strip().lower()
            if t:
                out.append(t)
    return out

def contains_token(text: str, tokens: List[str]) -> bool:
    if not text or not tokens:
        return False
    low = text.lower()
    return any(t in low for t in tokens)

def pe_signature_state(path: str) -> str:
    if pefile is None:
        return "N/A"
    try:
        pe = pefile.PE(path, fast_load=True)
        sec = pe.OPTIONAL_HEADER.DATA_DIRECTORY[pefile.DIRECTORY_ENTRY['IMAGE_DIRECTORY_ENTRY_SECURITY']]
        return "موقع" if getattr(sec, "VirtualAddress", 0) != 0 else "غير موقع"
    except Exception:
        return "N/A"

# ===== Multi Filter Proxy Model =====
class MultiFilterProxy(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.quickCriteria = {}  # dictionary: column index -> filter string

    def setQuickFilter(self, criteria: dict):
        self.quickCriteria = criteria
        self.invalidateFilter()

    def clearQuickFilter(self):
        self.quickCriteria = {}
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        # Apply quick filter criteria if any
        if self.quickCriteria:
            for col, fltr in self.quickCriteria.items():
                ix = self.sourceModel().index(source_row, col, source_parent)
                data = self.sourceModel().data(ix, Qt.DisplayRole)
                if fltr.lower() not in str(data).lower():
                    return False
        # Also apply the standard filtering mechanism
        return super().filterAcceptsRow(source_row, source_parent)

# ===== خيط الفحص =====
class ScannerThread(QThread):
    progress = pyqtSignal(int, int)
    finished = pyqtSignal(list, int, dict)
    error = pyqtSignal(str)
    def __init__(self, crit: Dict[str, Any]):
        super().__init__()
        self.crit = crit
        self._stop = False
    def stop(self):
        self._stop = True
    def _start_matches(self, raw: str) -> bool:
        f = self.crit.get("start_type", "all")
        if f == "all":
            return True
        return (raw or "").strip().lower() == f
    def run(self):
        # لجعل الأداة تعمل على ويندوز ولينوكس
        if os.name == "nt":
            try:
                services = list(psutil.win_service_iter())
            except Exception:
                services = []
        else:
            services = []
        try:
            procs = list(psutil.process_iter(['name','exe','username','pid','create_time','status']))
        except Exception as e:
            self.error.emit(f"Processes read failed: {e}")
            return
        scope = self.crit.get("scope", "both")
        total = (len(services) if scope in ("both", "services") else 0) + (len(procs) if scope in ("both", "processes") else 0)
        kws = split_tokens(self.crit.get("kws", []))
        use_age = self.crit.get("use_age", False)
        days = int(self.crit.get("days", 0))
        use_sig = self.crit.get("use_sig", False)
        startf = self.crit.get("start_type", "all")
        acct_filter = self.crit.get("account", "All")
        scanned = 0
        out = []
        # counters for indicators
        counters = {
            "procs_running": 0,
            "svcs_active": 0,
            "procs_kw": 0,
            "svcs_kw": 0,
            "procs_unsigned": 0
        }
        if scope in ("both", "services"):
            for s in services:
                if self._stop:
                    break
                scanned += 1
                self.progress.emit(scanned, total)
                try:
                    d = s.as_dict()
                except Exception:
                    continue
                name = d.get('name') or 'N/A'
                disp = d.get('display_name') or 'N/A'
                state = d.get('status') or 'N/A'
                start = (d.get('start_type') or '').lower()
                acc = d.get('username') or 'N/A'
                # فلترة الحساب
                if acct_filter.lower() != "all" and acc.lower() != acct_filter.lower():
                    continue
                exe = (d.get('binpath') or '').replace('"','').strip() or 'N/A'
                if startf != "all" and not self._start_matches(start):
                    continue
                reasons = []
                if kws and (contains_token(name, kws) or contains_token(disp, kws)):
                    reasons.append("Keyword match" if LANG!="ar" else "مطابقة كلمة مفتاحية")
                created = "N/A"
                if exe != "N/A" and os.path.isfile(exe):
                    try:
                        ctime = os.path.getctime(exe)
                        created = datetime.fromtimestamp(ctime).strftime("%Y-%m-%d %H:%M:%S")
                        if use_age and days > 0 and (datetime.now() - datetime.fromtimestamp(ctime)).days < days:
                            reasons.append("Recent item" if LANG!="ar" else "عنصر حديث")
                    except Exception:
                        pass
                sig = "N/A"
                if use_sig and exe != "N/A" and os.path.isfile(exe):
                    sst = pe_signature_state(exe)
                    sig = sst
                item = {
                    "type": "service",
                    "name": name,
                    "display_name": disp,
                    "state": state,
                    "start_type": start or "N/A",
                    "account": acc,
                    "exe": exe,
                    "sig": sig,
                    "reasons": reasons,
                    "created": created
                }
                out.append(item)
                if state.lower() == "running":
                    counters["svcs_active"] += 1
                if kws and reasons:
                    counters["svcs_kw"] += 1
        if scope in ("both", "processes"):
            for p in procs:
                if self._stop:
                    break
                scanned += 1
                self.progress.emit(scanned, total)
                try:
                    info = p.info
                except Exception:
                    continue
                name = info.get('name') or 'N/A'
                exe = info.get('exe') or 'N/A'
                acc = info.get('username') or 'N/A'
                # فلترة الحساب
                if acct_filter.lower() != "all" and acc.lower() != acct_filter.lower():
                    continue
                status = info.get('status') or 'N/A'
                reasons = []
                if kws and contains_token(name, kws):
                    reasons.append("Keyword match" if LANG!="ar" else "مطابقة كلمة مفتاحية")
                created = "N/A"
                if exe != "N/A" and os.path.isfile(exe):
                    try:
                        ctime = os.path.getctime(exe)
                        created = datetime.fromtimestamp(ctime).strftime("%Y-%m-%d %H:%M:%S")
                        if use_age and days > 0 and (datetime.now() - datetime.fromtimestamp(ctime)).days < days:
                            reasons.append("Recent item" if LANG!="ar" else "عنصر حديث")
                    except Exception:
                        pass
                sig = "N/A"
                if use_sig and exe != "N/A" and os.path.isfile(exe):
                    sst = pe_signature_state(exe)
                    sig = sst
                item = {
                    "type": "process",
                    "name": name,
                    "display_name": "N/A",
                    "state": status,
                    "start_type": "N/A",
                    "account": acc,
                    "exe": exe,
                    "sig": sig,
                    "reasons": reasons,
                    "created": created
                }
                out.append(item)
                if status.lower() == "running":
                    counters["procs_running"] += 1
                if kws and reasons:
                    counters["procs_kw"] += 1
                if use_sig and sig not in ("موقع", "signed", "N/A") and sig:
                    counters["procs_unsigned"] += 1
        self.finished.emit(out, total, counters)

# ===== كارد بسيط =====
class Card(QFrame):
    def __init__(self, title: str, clickable: bool = False):
        super().__init__()
        self.setObjectName("Card")
        self.setFrameShape(QFrame.StyledPanel)
        self.setFrameShadow(QFrame.Raised)
        self.v = QVBoxLayout(self)
        self.v.setContentsMargins(12,12,12,12)
        self.v.setSpacing(6)
        self.title_lbl = QLabel(title)
        self.title_lbl.setObjectName("CardTitle")
        self.value_lbl = QLabel("0")
        self.value_lbl.setObjectName("CardValue")
        self.v.addWidget(self.title_lbl)
        self.v.addWidget(self.value_lbl, 0, Qt.AlignRight)
        if clickable:
            self.setCursor(Qt.PointingHandCursor)

# ===== الواجهة =====
class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        self._load_keywords()
        self._build_ui()
        self._apply_theme_choice()
        self._apply_language()
    def _build_ui(self):
        self.setWindowTitle(tr("title"))
        self.resize(1500, 900)
        root = QWidget()
        self.setCentralWidget(root)
        root_v = QVBoxLayout(root)
        root_v.setContentsMargins(12,12,12,12)
        root_v.setSpacing(10)
        # أزرار أعلى + Exit + زر الخيارات
        top = QHBoxLayout()
        top.setSpacing(10)
        self.btn_scan = QPushButton(tr("scan"))
        self.btn_scan.setObjectName("ActionButton")
        self.btn_stop = QPushButton(tr("stop"))
        self.btn_stop.setObjectName("ActionButton")
        self.btn_stop.setEnabled(False)
        self.btn_clear = QPushButton(tr("clear"))
        self.btn_clear.setObjectName("ActionButton")
        self.btn_export = QPushButton(tr("export"))
        self.btn_export.setObjectName("ActionButton")
        self.btn_exit = QPushButton(tr("exit"))
        self.btn_exit.setObjectName("ActionButton")
        for b in (self.btn_scan, self.btn_stop, self.btn_clear, self.btn_export, self.btn_exit):
            b.setMinimumHeight(40)
        top.addWidget(self.btn_scan)
        top.addWidget(self.btn_stop)
        top.addWidget(self.btn_clear)
        top.addStretch(1)
        top.addWidget(self.btn_export)
        # زر خيارات عائم به قائمة منسدلة
        self.btn_options = QToolButton()
        self.btn_options.setText(tr("options"))
        self.btn_options.setPopupMode(QToolButton.InstantPopup)
        menu = QMenu()
        act_pdf = menu.addAction(tr("save_pdf"))
        act_compare = menu.addAction(tr("compare"))
        self.btn_options.setMenu(menu)
        act_pdf.triggered.connect(self._export_pdf)
        act_compare.triggered.connect(self._compare_results)
        top.addWidget(self.btn_options)
        top.addWidget(self.btn_exit)
        root_v.addLayout(top)
        # شريط الثيم/اللغة/النطاق (بدون شعار)
        header_bar = QHBoxLayout()
        header_bar.setSpacing(12)
        sel_box = QFrame()
        sel_box.setObjectName("MiniCard")
        sel_v = QHBoxLayout(sel_box)
        sel_v.setContentsMargins(12,12,12,12)
        sel_v.setSpacing(12)
        self.theme_combo = QComboBox()
        self.theme_combo.addItems([
            "غامق (Dark)", "فاتح (Light)",
            "Ocean Breeze", "Sunset Orange", "Midnight Purple",
            "Steel Gray", "Forest Green", "Ruby Red"
        ])
        self.lang_combo = QComboBox()
        self.lang_combo.addItems(["العربية", "English"])
        self.scope_combo = QComboBox()
        self.scope_combo.addItems([tr("scope_both"), tr("scope_services"), tr("scope_processes")])
        sel_v.addWidget(QLabel(tr("theme")))
        sel_v.addWidget(self.theme_combo)
        sel_v.addSpacing(20)
        sel_v.addWidget(QLabel(tr("lang")))
        sel_v.addWidget(self.lang_combo)
        sel_v.addSpacing(20)
        sel_v.addWidget(QLabel(tr("scope")))
        sel_v.addWidget(self.scope_combo)
        header_bar.addWidget(sel_box, 1)
        root_v.addLayout(header_bar)
        # الشبكة العليا: المعايير + الإحصاءات/الرسم
        upper = QGridLayout()
        upper.setHorizontalSpacing(10)
        upper.setVerticalSpacing(10)
        root_v.addLayout(upper, 2)
        # كارد المعايير
        self.card_crit = Card(tr("criteria"))
        cv = self.card_crit.v
        crit_grid = QGridLayout()
        crit_grid.setHorizontalSpacing(8)
        crit_grid.setVerticalSpacing(6)
        # الكروت المعدلة
        self.card_procs_running  = Card(tr_sub("cards", "procs_running"), clickable=True)
        self.card_svcs_active    = Card(tr_sub("cards", "svcs_active"), clickable=True)
        self.card_procs_kw       = Card(tr_sub("cards", "procs_kw"), clickable=True)
        self.card_svcs_kw        = Card(tr_sub("cards", "svcs_kw"), clickable=True)
        self.card_procs_unsigned = Card(tr_sub("cards", "procs_unsigned"), clickable=True)
        cards_grid = QGridLayout()
        cards_grid.setHorizontalSpacing(10)
        cards_grid.setVerticalSpacing(10)
        # الصف الأول: 3 كروت
        cards_grid.addWidget(self.card_procs_running, 0, 0)
        cards_grid.addWidget(self.card_svcs_active, 0, 1)
        cards_grid.addWidget(self.card_procs_kw, 0, 2)
        # الصف الثاني: 2 كروت
        cards_grid.addWidget(self.card_svcs_kw, 1, 0)
        cards_grid.addWidget(self.card_procs_unsigned, 1, 2)
        cards_grid.setColumnStretch(0, 1)
        cards_grid.setColumnStretch(1, 1)
        cards_grid.setColumnStretch(2, 1)
        cards_grid.setRowStretch(0, 1)
        cards_grid.setRowStretch(1, 1)
        crit_grid.addLayout(cards_grid, 0, 0, 1, 4)
        # قائمة الكلمات المفتاحية
        crit_grid.addWidget(QLabel(tr("keywords")), 1, 0, 1, 2)
        self.kws = QListWidget()
        btn_add_kw = QPushButton(tr("add"))
        btn_add_kw.setFixedWidth(36)
        btn_rem_kw = QPushButton(tr("remove"))
        btn_rem_kw.setFixedWidth(36)
        for k in getattr(self, "saved_kws", []):
            self.kws.addItem(QListWidgetItem(k))
        crit_grid.addWidget(self.kws, 2, 0, 1, 2)
        crit_grid.addWidget(btn_add_kw, 2, 2)
        crit_grid.addWidget(btn_rem_kw, 2, 3)
        # بدل زر حفظ وزر تحميل، نستخدم زر "حذف كل القائمة" في نفس الموقع
        self.btn_clear_kws = QPushButton(tr("delete_all"))
        crit_grid.addWidget(self.btn_clear_kws, 3, 0)
        # نوع البدء
        crit_grid.addWidget(QLabel(tr("start_type")), 4, 0)
        self.start_combo = QComboBox()
        self.start_combo.addItems([tr("start_all"), tr("start_auto"), tr("start_manual"), tr("start_disabled")])
        crit_grid.addWidget(self.start_combo, 4, 1, 1, 3)
        # الحساب (قائمة منسدلة)
        crit_grid.addWidget(QLabel(tr("acc")), 5, 0)
        self.acc_combo = QComboBox()
        self.acc_combo.addItems(["SYSTEM", "LocalSystem", "Administrator", "NetworkService", "LocalService", "Guest", "User", "All"])
        crit_grid.addWidget(self.acc_combo, 5, 1, 1, 3)
        # عمر / توقيع
        crit_grid.addWidget(QLabel(tr("days")), 6, 0)
        self.days_spin = QSpinBox()
        self.days_spin.setRange(0, 3650)
        self.days_spin.setValue(7)
        crit_grid.addWidget(self.days_spin, 6, 1)
        self.use_age = QCheckBox(tr("use_age"))
        crit_grid.addWidget(self.use_age, 6, 2)
        self.use_sig = QCheckBox(tr("use_sig"))
        crit_grid.addWidget(self.use_sig, 6, 3)
        cv.addLayout(crit_grid)
        upper.addWidget(self.card_crit, 0, 0)
        # كارد الإحصاءات/الرسم
        self.card_stats = Card(tr("stats"))
        sv = self.card_stats.v
        top_stats = QHBoxLayout()
        self.lbl_total = QLabel("0")
        self.lbl_susp = QLabel("0")
        self.lbl_rate = QLabel("0%")
        top_stats.addWidget(QLabel("Total/الإجمالي:"))
        top_stats.addWidget(self.lbl_total)
        top_stats.addStretch(1)
        top_stats.addWidget(QLabel("Matched/مطابق:"))
        top_stats.addWidget(self.lbl_susp)
        top_stats.addStretch(1)
        top_stats.addWidget(QLabel("Rate/النسبة:"))
        top_stats.addWidget(self.lbl_rate)
        sv.addLayout(top_stats)
        # مؤشرات الشيب (الجديدة)
        chips = QHBoxLayout()
        self.chip_procs_running = QLabel(tr_sub("chips", "procs_running") + ": 0")
        self.chip_procs_running.setObjectName("Chip")
        self.chip_svcs_active = QLabel(tr_sub("chips", "svcs_active") + ": 0")
        self.chip_svcs_active.setObjectName("Chip")
        chips.addWidget(self.chip_procs_running)
        chips.addWidget(self.chip_svcs_active)
        chips.addStretch(1)
        sv.addLayout(chips)
        sv.addWidget(QLabel(tr("reasons_chart")))
        if pg:
            self.plot = pg.PlotWidget()
            self.plot.setMinimumHeight(260)
            sv.addWidget(self.plot)
        else:
            self.plot = None
            sv.addWidget(QLabel("(Install pyqtgraph to see chart)"))
        upper.addWidget(self.card_stats, 0, 1)
        # النتائج: شريط فلترة + جدول
        self.card_results = Card(tr("results"))
        rv = self.card_results.v
        self.filter_bar = QFrame()
        self.filter_bar.setObjectName("MiniCard")
        self.filter_bar.setVisible(False)
        fb_h = QHBoxLayout(self.filter_bar)
        fb_h.setContentsMargins(5,5,5,5)
        fb_h.setSpacing(8)
        self.filter_col_combo = QComboBox()
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText(tr("search_bar_placeholder"))
        self.filter_apply_btn = QPushButton(tr("apply"))
        fb_h.addWidget(QLabel(tr("col")))
        fb_h.addWidget(self.filter_col_combo, 0)
        fb_h.addWidget(self.filter_edit, 1)
        fb_h.addWidget(self.filter_apply_btn, 0)
        rv.addWidget(QLabel(tr("filter_hint")))
        rv.addWidget(self.filter_bar)
        self.table = QTableView()
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(30)
        visible_rows = 6
        header_h = self.table.horizontalHeader().height()
        row_height = self.table.verticalHeader().defaultSectionSize()
        self.table.setMinimumHeight(header_h + row_height * visible_rows + 4)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.model = QStandardItemModel(0, len(table_headers()))
        self.model.setHorizontalHeaderLabels(table_headers())
        self.proxy = MultiFilterProxy(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.table.setModel(self.proxy)
        rv.addWidget(self.table)
        root_v.addWidget(self.card_results, 3)
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.status.addPermanentWidget(self.progress, 1)
        # إشارات
        btn_add_kw.clicked.connect(self._add_kw)
        btn_rem_kw.clicked.connect(self._rem_kw)
        self.btn_clear_kws.clicked.connect(self._clear_keywords)
        self.btn_scan.clicked.connect(self._start_scan)
        self.btn_stop.clicked.connect(self._stop_scan)
        self.btn_clear.clicked.connect(self._clear)
        self.btn_export.clicked.connect(self._export)
        self.btn_exit.clicked.connect(QApplication.instance().quit)
        self.theme_combo.currentIndexChanged.connect(self._apply_theme_choice)
        self.lang_combo.currentIndexChanged.connect(self._on_lang_change)
        self.filter_apply_btn.clicked.connect(self._apply_proxy_filter)
        self.filter_edit.returnPressed.connect(self._apply_proxy_filter)
        self.table.horizontalHeader().sectionClicked.connect(self._header_clicked)
        # تعديل أحداث النقر على الكروت بحيث يتم تطبيق فلتر سريع وفي حال عدم وجود نتائج يتم عرض رسالة تنبيه
        self.card_procs_running.mousePressEvent = lambda e: self._quick_filter("type:process;state:running")
        self.card_svcs_active.mousePressEvent = lambda e: self._quick_filter("type:service;state:running")
        self.card_procs_kw.mousePressEvent = lambda e: self._quick_filter("type:process;reason:keyword")
        self.card_svcs_kw.mousePressEvent = lambda e: self._quick_filter("type:service;reason:keyword")
        self.card_procs_unsigned.mousePressEvent = lambda e: self._quick_filter("type:process;sig:unsigned")
        self.shortcut_find = QKeySequence.Find
        self.installEventFilter(self)
        self.table.doubleClicked.connect(self._show_details)
        self.scanner = None
        self.last = []
        self._refresh_filter_columns()
        self._apply_card_colors()
    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress and QKeySequence(event.modifiers() | event.key()) == self.shortcut_find:
            self.filter_bar.setVisible(not self.filter_bar.isVisible())
            if self.filter_bar.isVisible():
                self.filter_edit.setFocus()
            return True
        return super().eventFilter(obj, event)
    def _on_lang_change(self, idx):
        global LANG
        LANG = "ar" if self.lang_combo.currentIndex() == 0 else "en"
        self._apply_language()
    def _apply_language(self):
        self.setWindowTitle(tr("title"))
        self.btn_scan.setText(tr("scan"))
        self.btn_stop.setText(tr("stop"))
        self.btn_clear.setText(tr("clear"))
        self.btn_export.setText(tr("export"))
        self.btn_exit.setText(tr("exit"))
        self.card_crit.title_lbl.setText(tr("criteria"))
        self.card_stats.title_lbl.setText(tr("stats"))
        self.card_results.title_lbl.setText(tr("results"))
        self.card_procs_running.title_lbl.setText(tr_sub("cards", "procs_running"))
        self.card_svcs_active.title_lbl.setText(tr_sub("cards", "svcs_active"))
        self.card_procs_kw.title_lbl.setText(tr_sub("cards", "procs_kw"))
        self.card_svcs_kw.title_lbl.setText(tr_sub("cards", "svcs_kw"))
        self.card_procs_unsigned.title_lbl.setText(tr_sub("cards", "procs_unsigned"))
        self.filter_edit.setPlaceholderText(tr("search_bar_placeholder"))
        self.filter_apply_btn.setText(tr("apply"))
        self.model.setHorizontalHeaderLabels(table_headers())
        self._refresh_filter_columns()
    def _refresh_filter_columns(self):
        self.filter_col_combo.blockSignals(True)
        self.filter_col_combo.clear()
        for i in range(self.model.columnCount()):
            self.filter_col_combo.addItem(self.model.headerData(i, Qt.Horizontal))
        self.filter_col_combo.blockSignals(False)
    def _header_clicked(self, index: int):
        self.filter_bar.setVisible(True)
        self.filter_col_combo.setCurrentIndex(index)
        self.filter_edit.setFocus()
    def _apply_proxy_filter(self):
        # عند استخدام شريط الفلترة، يتم مسح فلتر الكيويك
        self.proxy.clearQuickFilter()
        col = self.filter_col_combo.currentIndex()
        text = self.filter_edit.text().strip()
        self.proxy.setFilterKeyColumn(col)
        self.proxy.setFilterFixedString(text)
        if self.proxy.rowCount() == 0:
            self.status.showMessage("لا توجد نتائج مطابقة")
    def _quick_filter(self, expr: str):
        # تحسين دالة الفلترة السريعة بحيث تُظهر رسالة تنبيه في حال عدم وجود نتائج مطابقة
        parts = {}
        for p in expr.split(";"):
            if ":" in p:
                key, value = p.split(":", 1)
                parts[key] = value
        headers = [self.model.headerData(i, Qt.Horizontal) for i in range(self.model.columnCount())]
        def col_index(name_ar, name_en):
            try:
                return headers.index(name_ar) if LANG=="ar" else headers.index(name_en)
            except ValueError:
                return -1
        criteria = {}
        if "type" in parts:
            idx = col_index("النوع", "Type")
            if idx >= 0:
                criteria[idx] = tr("type_process") if parts["type"]=="process" else tr("type_service")
        if "state" in parts:
            idx = col_index("الحالة", "State")
            if idx >= 0:
                criteria[idx] = "running" if LANG!="ar" else "جارية"
        if "reason" in parts:
            idx = col_index("أسباب", "Reasons")
            if idx >= 0:
                criteria[idx] = "Keyword" if LANG!="ar" else "مطابقة كلمة مفتاحية"
        if "sig" in parts:
            idx = col_index("Signature", "Signature")
            if idx >= 0:
                # لضبط كرت العمليات غير المُوقَّعة، نقارن بعرض "Invalid"
                criteria[idx] = "Invalid"
        self.proxy.setQuickFilter(criteria)
        if self.proxy.rowCount() == 0:
            QMessageBox.information(self, tr("title"), "لا توجد نتائج مطابقة")
        if not self.filter_bar.isVisible():
            self.filter_bar.setVisible(True)
    def _show_details(self, index: QModelIndex):
        if not index.isValid():
            return
        r = self.proxy.mapToSource(index).row()
        data = {}
        for c in range(self.model.columnCount()):
            key = self.model.headerData(c, Qt.Horizontal)
            val = self.model.item(r, c).text() if self.model.item(r, c) else ""
            data[str(key)] = val
        lines = [f"{k}: {v}" for k, v in data.items()]
        QMessageBox.information(self, tr("title"), "\n".join(lines))
    def _add_kw(self):
        t, ok = QInputDialog.getText(self, tr("keywords"), tr("add"))
        if ok and t.strip():
            vals = [self.kws.item(i).text() for i in range(self.kws.count())]
            if t.strip() not in vals:
                self.kws.addItem(QListWidgetItem(t.strip()))
                self._save_keywords()
    def _rem_kw(self):
        r = self.kws.currentRow()
        if r >= 0:
            self.kws.takeItem(r)
            self._save_keywords()
    def _clear_keywords(self):
        self.kws.clear()
        self._save_keywords()
    def _save_keywords(self):
        data = {"kws": [self.kws.item(i).text() for i in range(self.kws.count())]}
        try:
            with open(LISTS_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
    def _load_keywords(self):
        if LISTS_FILE.exists():
            try:
                data = json.load(open(LISTS_FILE, "r", encoding="utf-8"))
                self.saved_kws = data.get("kws", ["ssh", "reverse", "backdoor"])
            except Exception:
                self.saved_kws = ["ssh", "reverse", "backdoor"]
        else:
            self.saved_kws = ["ssh", "reverse", "backdoor"]
    def _criteria(self) -> Dict[str, Any]:
        start_text = self.start_combo.currentText()
        start_token = "all"
        if start_text in (tr("start_auto"),):
            start_token = "automatic"
        elif start_text in (tr("start_manual"),):
            start_token = "manual"
        elif start_text in (tr("start_disabled"),):
            start_token = "disabled"
        scope_idx = self.scope_combo.currentIndex()
        scope = "both" if scope_idx == 0 else ("services" if scope_idx == 1 else "processes")
        account = self.acc_combo.currentText()
        return {
            "kws": [self.kws.item(i).text() for i in range(self.kws.count())],
            "days": self.days_spin.value(),
            "use_age": self.use_age.isChecked(),
            "use_sig": self.use_sig.isChecked(),
            "start_type": start_token,
            "scope": scope,
            "account": account
        }
    def _start_scan(self):
        crit = self._criteria()
        self._clear_table_only()
        self.last = []
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.btn_scan.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.scanner = ScannerThread(crit)
        self.scanner.progress.connect(self._on_progress)
        self.scanner.finished.connect(self._on_finished)
        self.scanner.error.connect(self._on_error)
        self.scanner.start()
    def _stop_scan(self):
        if self.scanner:
            self.scanner.stop()
        self.btn_scan.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)
    def _clear(self):
        self._clear_table_only()
        self.last = []
        self.lbl_total.setText("0")
        self.lbl_susp.setText("0")
        self.lbl_rate.setText("0%")
        self.chip_procs_running.setText(tr_sub("chips", "procs_running") + ": 0")
        self.chip_svcs_active.setText(tr_sub("chips", "svcs_active") + ": 0")
        if self.plot:
            self.plot.clear()
    def _clear_table_only(self):
        self.model.removeRows(0, self.model.rowCount())
        self.status.clearMessage()
    def _on_progress(self, cur: int, total: int):
        self.progress.setMaximum(total)
        self.progress.setValue(cur)
        self.status.showMessage(tr("progress").format(cur, total))
        self.lbl_total.setText(str(total))
    def _on_finished(self, items: List[Dict[str, Any]], total: int, counters: dict):
        self.last = items
        self.model.setRowCount(0)
        headers = tr("table_headers")
        for it in items:
            # تعديل عرض عمود Signature: عرض "Valid" أو "Invalid"
            sig_val = it["sig"]
            if sig_val == "N/A":
                sig_display = "N/A"
            elif sig_val in ("موقع", "signed"):
                sig_display = "Valid"
            else:
                sig_display = "Invalid"
            row = [
                (tr("type_service") if it["type"]=="service" else tr("type_process")),
                it["name"], it["display_name"], it["state"], it["start_type"],
                it["account"], it["exe"], sig_display,
                ", ".join(it["reasons"]), it["created"]
            ]
            r = self.model.rowCount()
            self.model.insertRow(r)
            for c, val in enumerate(row):
                item = QStandardItem(str(val))
                item.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter)
                self.model.setItem(r, c, item)
        self._refresh_filter_columns()
        self.btn_scan.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)
        self.lbl_total.setText(str(total))
        self.lbl_susp.setText(str(len(items)))
        rate = (len(items) / total * 100) if total > 0 else 0
        self.lbl_rate.setText(f"{rate:.2f}%")
        self.status.showMessage(tr("done").format(len(items), total))
        self.card_procs_running.value_lbl.setText(str(counters.get("procs_running", 0)))
        self.card_svcs_active.value_lbl.setText(str(counters.get("svcs_active", 0)))
        self.card_procs_kw.value_lbl.setText(str(counters.get("procs_kw", 0)))
        self.card_svcs_kw.value_lbl.setText(str(counters.get("svcs_kw", 0)))
        self.card_procs_unsigned.value_lbl.setText(str(counters.get("procs_unsigned", 0)))
        self.chip_procs_running.setText(tr_sub("chips", "procs_running") + f": {counters.get('procs_running', 0)}")
        self.chip_svcs_active.setText(tr_sub("chips", "svcs_active") + f": {counters.get('svcs_active', 0)}")
        if self.plot:
            self.plot.clear()
            counts: Dict[str, int] = {}
            for it in items:
                for r in it.get("reasons", []):
                    counts[r] = counts.get(r, 0) + 1
            palette = [
                (58,134,255), (138,43,226), (255,99,132),
                (255,159,64), (75,192,192), (153,102,255),
                (255,205,86), (100,255,100)
            ]
            xs = list(range(len(counts)))
            names = list(counts.keys())
            ys = [counts[k] for k in names]
            for i, (x, h) in enumerate(zip(xs, ys)):
                color = palette[i % len(palette)]
                bar = pg.BarGraphItem(x=[x], height=[h], width=0.7, brush=color, pen=(230,230,230))
                self.plot.addItem(bar)
            extra_indicators = [("Running Processes", counters.get("procs_running", 0)),
                                ("Active Services", counters.get("svcs_active", 0)),
                                ("Invalid Signature", counters.get("procs_unsigned", 0))]
            start_x = len(counts)
            extra_palette = [(50,205,50), (30,144,255), (255,69,0)]
            for i, (label, value) in enumerate(extra_indicators):
                bar = pg.BarGraphItem(x=[start_x + i], height=[value], width=0.7, brush=extra_palette[i], pen=(230,230,230))
                self.plot.addItem(bar)
                text = pg.TextItem(text=label, color='w', anchor=(0.5, -0.5))
                text.setPos(start_x + i, value)
                self.plot.addItem(text)
            ax = self.plot.getAxis('bottom')
            all_labels = names + [x[0] for x in extra_indicators]
            ticks = [(i, all_labels[i]) for i in range(len(all_labels))]
            ax.setTicks([ticks])
            self.plot.showGrid(x=True, y=True, alpha=0.3)
        # حفظ النتائج الجديدة في ملف JSON بعد الفحص
        self._save_results(self.last)
    def _on_error(self, msg: str):
        QMessageBox.warning(self, "Error", msg)
        self.btn_scan.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)
    def _export(self):
        if not self.last:
            QMessageBox.information(self, tr("title"), "لا نتائج للتصدير." if LANG=="ar" else "No results to export.")
            return
        if openpyxl is None:
            QMessageBox.warning(self, "تنبيه" if LANG=="ar" else "Note", tr("need_openpyxl"))
            return
        fname, _ = QFileDialog.getSaveFileName(self, "حفظ Excel" if LANG=="ar" else "Save Excel", str(Path.home()), "Excel (*.xlsx)")
        if not fname:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ScanResults"
            headers = tr("table_headers")
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)
            for r, it in enumerate(self.last, start=2):
                sig_val = it["sig"]
                if sig_val == "N/A":
                    sig_display = "N/A"
                elif sig_val in ("موقع", "signed"):
                    sig_display = "Valid"
                else:
                    sig_display = "Invalid"
                row = [
                    (tr("type_service") if it["type"]=="service" else tr("type_process")),
                    it["name"], it["display_name"], it["state"], it["start_type"],
                    it["account"], it["exe"], sig_display,
                    "; ".join(it["reasons"]), it["created"]
                ]
                for c, v in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=v)
            for i, _ in enumerate(headers, start=1):
                col = get_column_letter(i)
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col])
                ws.column_dimensions[col].width = min(max_len + 4, 80)
            wb.save(fname)
            QMessageBox.information(self, tr("title"), f"{'تم الحفظ' if LANG=='ar' else 'Saved'}: {fname}")
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
    def _export_pdf(self):
        if not self.last:
            QMessageBox.information(self, tr("title"), "لا نتائج للتصدير.")
            return
        fname, _ = QFileDialog.getSaveFileName(self, "حفظ PDF", str(Path.home()), "PDF (*.pdf)")
        if not fname:
            return
        try:
            c = canvas.Canvas(fname, pagesize=letter)
            width, height = letter
            margin = 50
            # عنوان التقرير
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, height - margin, "نتائج الفحص:")
            # إذا كان المخطط موجوداً، نقوم بالتقاط صورة الرسم البياني وإدراجها في التقرير
            image_height = 200
            image_width = width - (2 * margin)
            if self.plot:
                temp_image_path = "temp_plot.png"
                pixmap = self.plot.grab()
                pixmap.save(temp_image_path, "PNG")
                # وضع الصورة في التقرير
                c.drawImage(temp_image_path, margin, height - margin - image_height - 20, width=image_width, height=image_height)
                os.remove(temp_image_path)
                text_start_y = height - margin - image_height - 40
            else:
                text_start_y = height - margin - 20
            # إعداد النصوص للبيانات
            c.setFont("Helvetica", 10)
            y_position = text_start_y
            line_height = 14
            for item in self.last:
                sig_val = item["sig"]
                if sig_val == "N/A":
                    sig_display = "N/A"
                elif sig_val in ("موقع", "signed"):
                    sig_display = "Valid"
                else:
                    sig_display = "Invalid"
                line = f"النوع: {item['type']} | الاسم: {item['name']} | الحالة: {item['state']} | Signature: {sig_display}"
                c.drawString(margin, y_position, line)
                y_position -= line_height
                if y_position < margin:
                    c.showPage()
                    y_position = height - margin
                    c.setFont("Helvetica", 10)
            c.save()
            QMessageBox.information(self, tr("title"), f"تم حفظ التقرير كـ PDF: {fname}")
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
    def _save_results(self, results):
        # حفظ النتائج الحالية في ملف JSON
        try:
            with open("scan_results.json", "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))
    def _load_previous_results(self):
        # تحميل النتائج السابقة من ملف JSON
        try:
            with open("scan_results.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return []
    def _compare_results(self):
        previous_results = self._load_previous_results()
        current_results = self.last
        differences = []
        for current in current_results:
            if current not in previous_results:
                differences.append(current)
        if differences:
            QMessageBox.information(self, "مقارنة النتائج", f"تم العثور على {len(differences)} اختلافات.")
        else:
            QMessageBox.information(self, "مقارنة النتائج", "لا توجد اختلافات بين النتائج.")
    def _apply_card_colors(self):
        self.card_procs_running.setStyleSheet("background-color: #4f8ef7; color: white; border-radius: 10px;")
        self.card_svcs_active.setStyleSheet("background-color: #52c41a; color: white; border-radius: 10px;")
        self.card_procs_kw.setStyleSheet("background-color: #fa8c16; color: white; border-radius: 10px;")
        self.card_svcs_kw.setStyleSheet("background-color: #722ed1; color: white; border-radius: 10px;")
        self.card_procs_unsigned.setStyleSheet("background-color: #f5222d; color: white; border-radius: 10px;")
    def _apply_theme_choice(self):
        idx = self.theme_combo.currentIndex()
        if idx == 0:
            self._apply_dark()
        elif idx == 1:
            self._apply_light()
        elif idx == 2:
            self._apply_ocean_breeze()
        elif idx == 3:
            self._apply_sunset_orange()
        elif idx == 4:
            self._apply_midnight_purple()
        elif idx == 5:
            self._apply_steel_gray()
        elif idx == 6:
            self._apply_forest_green()
        elif idx == 7:
            self._apply_ruby_red()
    def _base_style(self, fg, bg, card_bg, border, header_bg, header_fg, sel_bg, chip_bg, chip_fg, accent="#4f6fff"):
        self.setStyleSheet(f"""
        QWidget {{ font-family: "Segoe UI", Tahoma; font-size: 8pt; color: {fg}; }}
        QMainWindow {{ background: {bg}; }}
        #Card {{
            background: {card_bg};
            border: 1px solid {border}; border-radius: 14px;
        }}
        #CardTitle {{ font-weight:600; font-size: 12pt; color:{header_fg}; }}
        #CardValue {{ font-weight:700; font-size: 8pt; }}
        #MiniCard {{
            background:{card_bg};
            border:1px solid {border};
            border-radius: 12px;
        }}
        QLabel {{ color:{fg}; }}
        QLineEdit, QComboBox, QListWidget, QSpinBox, QAbstractItemView {{
            background:{bg}; border:1px solid {border}; border-radius:8px; padding:4px; color:{fg};
        }}
        QComboBox::drop-down {{ border:0px; }}
        QCheckBox {{ spacing:6px; }}
        QTableView {{
            background:{bg}; gridline-color:{border}; border:1px solid {border}; border-radius:12px;
        }}
        QHeaderView::section {{
            background:{header_bg}; color:{header_fg}; border:0px; padding:8px; border-right:1px solid {border};
        }}
        QTableView::item:selected {{ background:{sel_bg}; }}
        QStatusBar {{ background:{card_bg}; border-top:1px solid {border}; }}
        QProgressBar {{
            background:{bg}; border:1px solid {border}; border-radius:8px; text-align:center; color:{fg};
        }}
        QProgressBar::chunk {{ background:{accent}; border-radius:8px; }}
        QPushButton#ActionButton {{
            background: {accent}; color:white; border:0; border-radius: 10px; padding:10px 16px; font-weight:600;
        }}
        QPushButton#ActionButton:disabled {{ background:#808080; color:#e0e0e0; }}
        QPushButton {{
            background:{card_bg}; border:1px solid {border}; border-radius:8px; padding:6px 10px; color:{fg};
        }}
        QPushButton:hover {{ filter: brightness(115%); }}
        #Chip {{
            background:{chip_bg}; color:{chip_fg}; border:1px solid {border}; border-radius:12px; padding:4px 10px;
        }}
        QMessageBox {{
            background: {card_bg};
            color: {fg};
        }}
        """)
        if self.plot:
            self.plot.setBackground(bg)
            self.plot.getAxis('left').setPen(header_fg)
            self.plot.getAxis('bottom').setPen(header_fg)
            self.plot.getAxis('left').setTextPen(header_fg)
            self.plot.getAxis('bottom').setTextPen(header_fg)
    def _apply_dark(self):
        self._base_style(
            fg="#e6e6e6", bg="#0f1220", card_bg="#1b2034", border="#2f3b5a",
            header_bg="#172038", header_fg="#a8c1ff", sel_bg="#23345d",
            chip_bg="#23345d", chip_fg="#e6e6e6", accent="#3941a5"
        )
    def _apply_light(self):
        self._base_style(
            fg="#1c2030", bg="#f5f7fb", card_bg="#ffffff", border="#e3e8f5",
            header_bg="#eef2ff", header_fg="#3b5bcc", sel_bg="#dfe6ff",
            chip_bg="#eef2ff", chip_fg="#2d3c77", accent="#4f6fff"
        )
    def _apply_ocean_breeze(self):
        self._base_style(
            fg="#eaf6ff", bg="#0a1e2e", card_bg="#11354b", border="#1e4e6a",
            header_bg="#15445d", header_fg="#9ad8ff", sel_bg="#1a5674",
            chip_bg="#15445d", chip_fg="#eaf6ff", accent="#2aa1d3"
        )
    def _apply_sunset_orange(self):
        self._base_style(
            fg="#2b1100", bg="#fff6f0", card_bg="#fff0e6", border="#ffd6bf",
            header_bg="#ffe3d2", header_fg="#a13a00", sel_bg="#ffd6bf",
            chip_bg="#ffe3d2", chip_fg="#5a2300", accent="#ff7a33"
        )
    def _apply_midnight_purple(self):
        self._base_style(
            fg="#f0eaff", bg="#120a1e", card_bg="#1e1230", border="#3a2b5a",
            header_bg="#2a1946", header_fg="#cbb3ff", sel_bg="#3a2b5a",
            chip_bg="#2a1946", chip_fg="#f0eaff", accent="#8a2be2"
        )
    def _apply_steel_gray(self):
        self._base_style(
            fg="#e9edf2", bg="#1f242b", card_bg="#2a313a", border="#3a444f",
            header_bg="#343e4a", header_fg="#cfe3ff", sel_bg="#3a444f",
            chip_bg="#343e4a", chip_fg="#e9edf2", accent="#5a8dee"
        )
    def _apply_forest_green(self):
        self._base_style(
            fg="#e9ffe9", bg="#0f1f14", card_bg="#163522", border="#224a31",
            header_bg="#1b4029", header_fg="#a8ffbf", sel_bg="#235a37",
            chip_bg="#1b4029", chip_fg="#e9ffe9", accent="#2fbf71"
        )
    def _apply_ruby_red(self):
        self._base_style(
            fg="#fff0f3", bg="#2b060c", card_bg="#3e0a12", border="#5a101c",
            header_bg="#4a0c16", header_fg="#ffb3c1", sel_bg="#5a101c",
            chip_bg="#4a0c16", chip_fg="#fff0f3", accent="#e63961"
        )
def main():
    global LANG
    LANG = "ar"
    app = QApplication(sys.argv)
    w = Main()
    w.show()
    sys.exit(app.exec_())
if __name__ == "__main__":
    main()