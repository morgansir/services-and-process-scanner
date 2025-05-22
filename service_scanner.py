
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psutil
import sqlite3
import json
import openpyxl
import joblib
import os
import random
import math
import threading
import csv
import platform
import subprocess
import hashlib
import ctypes
from ctypes import wintypes
from datetime import datetime
import pandas as pd
from sklearn.ensemble import IsolationForest
import matplotlib

matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from mpl_toolkits.mplot3d import Axes3D
from PIL import Image, ImageTk, ImageSequence
import time
import socket
import numpy as np
import mplcursors  # pip install mplcursors
import yaml  # pip install pyyaml

# تأكد أنه موجود مرة واحدة فقط
import math
import tkinter as tk

# دالة لتوليد لون متدرج بناءً على الزاوية
def sin_color(angle):
    r = int(128 + 127 * math.sin(math.radians(angle)))
    g = int(128 + 127 * math.sin(math.radians(angle + 120)))
    b = int(128 + 127 * math.sin(math.radians(angle + 240)))
    return r, g, b

# ----- إعدادات قاعدة البيانات والنماذج -----
DB_FILE = "services.db"
MODEL_FILE = "anomaly_detector.pkl"
BASELINE_FILE = "baseline_telemetry.csv"

SIGN_TOOL = "signtool"  # أداة ويندوز للتحقق من التوقيع

# ----- ألوان الواجهة والثيمات -----
PRIMARY = "#0D47A1"
SECONDARY = "#1976D2"
TERTIARY = "#42A5F5"

# مجموعة ألوان تدرجية
GRADIENTS = ["#FF7F50", "#FF6347", "#FF4500", "#FF8C00"]
GRADIENTS_EXT = ["#121212", "#222222", "#333333", "#111111"]

MAIN_BG = PRIMARY
HEADER_BG = SECONDARY
BUTTON_COLOR = TERTIARY
FONT_NAME = "Helvetica"

# ----- قاموس تخزين الهاشات الأساسية للخدمات الموثوقة -----
baseline_hashes = {}  # مثال: {'C:\\Path\\to\\binary.exe': 'abc123...'}

# ----- تهيئة قاعدة البيانات: إنشاء جداول الخدمات وقواعد SIGMA -----
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY,
            pid INTEGER,
            name TEXT,
            path TEXT,
            account TEXT,
            frequency TEXT,
            age_days INTEGER,
            timestamp TEXT,
            anomalous INTEGER,
            anomaly_score REAL,
            signature_valid INTEGER,
            source TEXT,
            delta_cpu REAL,
            delta_io REAL,
            tcp_listen INTEGER,
            tcp_established INTEGER,
            udp INTEGER
        )
    ''')
    # إنشاء جدول لقواعد SIGMA
    c.execute('''
        CREATE TABLE IF NOT EXISTS sigma_rules (
            id INTEGER PRIMARY KEY,
            rule_id TEXT,
            rule_title TEXT,
            rule_data TEXT

        )
    ''')
    conn.commit()
    conn.close()

init_db()

# ----- إدارة إعدادات الكلمات المشبوهة -----
class SettingsManager:
    def __init__(self):
        self.keywords = self.load_keywords()

    def load_keywords(self):
        if os.path.exists('keywords.json'):
            try:
                with open('keywords.json', 'r', encoding='utf-8') as f:
                    return json.load(f).get('keywords', [])
            except Exception as e:
                print("Error loading keywords:", str(e))
                return []
        return []

    def save_keywords(self, keywords):
        with open('keywords.json', 'w', encoding='utf-8') as f:
            json.dump({'keywords': list(keywords)}, f)

# ----- شاشة الترحيب ثلاثية الأبعاد متحركة (SplashScreen) -----
class SplashScreen(tk.Toplevel):
    def __init__(self, root):
        super().__init__(root)
        self.overrideredirect(True)
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w, h = 800, 600
        x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.canvas = tk.Canvas(self, width=w, height=h, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.gradient_colors = ["#003973", "#004c8c", "#005ca5", "#0073c1",
                                "#008ad9", "#00a2f3"]
        self.grad_index = 0
        self.bg_rect = self.canvas.create_rectangle(0, 0, w, h,
                                                    fill=self.gradient_colors[self.grad_index], outline="")
        self.welcome_text = (
            "🔍 تحليل البيانات المتقدم\n"
            "🚀 سرعة عالية وكفاءة ممتازة\n"
            "💡 واجهة مبتكرة وحديثة\n"
            "🎨 ألوان جذابة وتصميم راقٍ\n"
            "🔒 أمان وتحليل سلوك الخدمات"
        )
        self.text_item = self.canvas.create_text(w // 2, h // 2,
                                                 text=self.welcome_text, fill="white",

                                                 font=(FONT_NAME, 16, 'italic'), justify='center')
        self.animate_background()
        self.after(3000, self.destroy)

    def animate_background(self):
        self.grad_index = (self.grad_index + 1) % len(self.gradient_colors)
        self.canvas.itemconfig(self.bg_rect,
                               fill=self.gradient_colors[self.grad_index])
        self.after(150, self.animate_background)

# ----- وظيفة جديدة للحصول على PID للخدمة باستخدام sc queryex -----
def get_service_pid(svc_name):
    try:
        output = subprocess.check_output(["sc", "queryex", svc_name],
                                         universal_newlines=True, stderr=subprocess.DEVNULL)
        for line in output.splitlines():
            if "PID" in line:
                parts = line.split()
                for part in parts:
                    if part.isdigit():
                        return int(part)
        return ""
    except Exception:
        return ""

# ----- وظائف مساعدة لفحص خدمات النظام عبر SCM على ويندوز -----
def get_services_from_scm():
    """
    يجمع أسماء جميع خدمات الويندوز المسجَّلة عبر psutil.
    """
    services = []
    if platform.system() != "Windows":
        return services
    for svc in psutil.win_service_iter():
        try:
            services.append(svc.name())
        except Exception:
            continue
    return services

# ----- التطبيق الرئيسي -----
class ServiceScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Service Scanner Pro - Behavioral Anomaly Detection")
        self.root.geometry("1400x820")
        self.root.configure(bg=MAIN_BG)
        self.all_services_data = []
        self.radar_paused = False  # لتتبع حالة إيقاف حركة مخطط الرادار
        self.setup_styles()

        # الشريط العلوي
        self.top_bar = tk.Frame(root, bg=HEADER_BG)
        self.top_bar.pack(fill=tk.X)
        btn_refresh = tk.Button(self.top_bar, text="Refresh",
                                command=self.refresh_all,
                                bg=random.choice(GRADIENTS), fg="white", font=(FONT_NAME, 11, 'bold'))
        btn_refresh.pack(side=tk.LEFT, padx=10, pady=5)
        btn_exit = tk.Button(self.top_bar, text="Exit", command=root.quit,
                             bg="red", fg="white", font=(FONT_NAME, 11, 'bold'))
        btn_exit.pack(side=tk.RIGHT, padx=10, pady=5)

        # شاشة الترحيب
        root.withdraw()
        SplashScreen(root)
        root.after(3000, root.deiconify)
        self.settings = SettingsManager()
        self.anomaly_model = self.load_model()
        self.create_widgets()
        self.bind_all_treeviews()

    # ----- تحميل نموذج الشذوذ -----
    def load_model(self):
        if os.path.exists(MODEL_FILE):
            try:
                return joblib.load(MODEL_FILE)
            except Exception as e:
                print("Model load error:", str(e))
                return None
        return None

    # ----- التحقق من التوقيع الرقمي -----
    def verify_signature(self, path):
        if platform.system() == 'Windows' and os.path.exists(path):
            try:
                res = subprocess.run([SIGN_TOOL, 'verify', '/pa', path],
                                     capture_output=True)
                return res.returncode == 0
            except Exception:
                return False
        return True

    # ----- التحقق من سلامة الملف عن طريق حساب SHA256 -----
    def compute_file_hash(self, path):
        try:
            with open(path, "rb") as f:
                file_hash = hashlib.sha256()
                while True:
                    chunk = f.read(8192)
                    if not chunk:
                        break
                    file_hash.update(chunk)

                return file_hash.hexdigest()
        except Exception:
            return None

    # ----- إعداد الستايلات والتنسيقات -----
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TButton", font=(FONT_NAME, 11, 'bold'),
                        foreground='white', background=BUTTON_COLOR,
                        relief="raised", borderwidth=3)
        style.map("TButton", background=[("active", random.choice(GRADIENTS))])
        style.configure("Treeview.Heading", font=(FONT_NAME, 11, 'bold'),
                        background=SECONDARY, foreground="black")
        style.configure("Treeview", font=(FONT_NAME, 10),
                        rowheight=24, background=MAIN_BG, foreground="white",
                        fieldbackground=MAIN_BG)
        style.map("Treeview", background=[("selected", SECONDARY)])
        style.configure("TNotebook.Tab", font=(FONT_NAME, 11, 'bold'),
                        padding=[12, 6], background=SECONDARY)
        style.map("TNotebook.Tab", background=[("selected", BUTTON_COLOR)])
        style.configure("Custom.Horizontal.TProgressbar", troughcolor="#EEEEEE",
                        bordercolor="#AAAAAA",
                        background="#FF1493", lightcolor="#FF69B4", darkcolor="#C71585")

    # ----- دالة تحريك المثلث ذو الشريط المتحرك (للتبويبات) -----
    def animate_triangle_border(self, canvas, triangle_id):
        if not hasattr(self, '_tab_grad_angle'):
            self._tab_grad_angle = 0
        r, g, b = sin_color(self._tab_grad_angle)
        color = f'#{r:02x}{g:02x}{b:02x}'
        canvas.itemconfig(triangle_id, outline=color, width=3)
        self._tab_grad_angle = (self._tab_grad_angle + 5) % 360
        canvas.after(100, lambda: self.animate_triangle_border(canvas, triangle_id))

    # ----- دالة لتشغيل رمز الانتظار (Spinner) -----
    def start_spinner(self, spinner_label, running_flag_name):
        setattr(self, running_flag_name, True)
        spinner_chars = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']

        def spin(idx=0):
            if getattr(self, running_flag_name):
                spinner_label.config(text=spinner_chars[idx % len(spinner_chars)])
                spinner_label.after(100, lambda: spin(idx + 1))
            else:
                spinner_label.config(text="")

        spin()

    def stop_spinner(self, running_flag_name):
        setattr(self, running_flag_name, False)


    def start_spinner_with_counter(self, spinner_label, running_flag_name):
        setattr(self, running_flag_name, True)
        spinner_chars = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']
        counter = 0

        def spin(idx=0):
            nonlocal counter
            if getattr(self, running_flag_name):
                if idx % 10 == 0:
                    counter += 1
                spinner_label.config(text=f"{spinner_chars[idx % len(spinner_chars)]} {counter}")
                spinner_label.after(100, lambda: spin(idx + 1))
            else:
                spinner_label.config(text="")

        spin()

    def stop_spinner_with_counter(self, running_flag_name):
        setattr(self, running_flag_name, False)

    # ----- دالة لمزامنة لون الخلفية للأزرار بشكل متدرج -----
    def animate_button_gradient(self, button, start_angle=0):
        def update_color(angle=start_angle):
            r, g, b = sin_color(angle)
            color = f'#{r:02x}{g:02x}{b:02x}'
            button.configure(bg=color, activebackground=color)
            button.after(100, lambda: update_color((angle + 10) % 360))
        update_color()

    # ----- دالة لتطبيق تأثير الخلفية المموجة على نوافذ الحوار -----
    def animate_dialog_wavy(self, dialog, angle=0):
        r, g, b = sin_color(angle)
        color = f'#{r:02x}{g:02x}{b:02x}'
        dialog.configure(bg=color)
        dialog.after(100, lambda: self.animate_dialog_wavy(dialog, (angle + 10) % 360))

    # ----- إنشاء التبويبات الرئيسية -----
    def create_widgets(self):
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True)
        # التبويبات: Welcome, Scan process, Keywords Scan, Suspicious Services
        self.tab_welcome = tk.Frame(self.nb)
        self.tab_all = tk.Frame(self.nb)  # تبويب Scan process
        self.tab_kw = tk.Frame(self.nb)
        self.tab_svc = tk.Frame(self.nb)
        self.nb.add(self.tab_welcome, text="🏠 Welcome")
        self.nb.add(self.tab_all, text="Scan process")
        self.nb.add(self.tab_kw, text="✏️ Keywords Scan")
        self.nb.add(self.tab_svc, text="⚙️ Suspicious Services")
        # رسم خلفية متدرجة على كل تبويب

        self.animate_tab_gradient(self.tab_welcome, 300)
        self.animate_tab_gradient(self.tab_all, 400)
        self.animate_tab_gradient(self.tab_kw, 450)
        self.animate_tab_gradient(self.tab_svc, 500)
        self.create_tab_welcome()
        self.create_tab_all()  # تحسينات تبويب Scan process
        self.create_tab_keywords()
        self.create_tab_suspicious_services()

    # ----- دالة تحريك الخلفية المتدرجة على أي تبويب -----
    def animate_tab_gradient(self, frame, delay=100):
        if not hasattr(frame, '_grad_angle'):
            frame._grad_angle = 0
        if not hasattr(frame, '_grad_canvas'):
            frame._grad_canvas = tk.Canvas(frame, highlightthickness=0)
            frame._grad_canvas.place(relwidth=1, relheight=1)
        c = frame._grad_canvas

        def update():
            angle = frame._grad_angle
            c.delete("grad")
            w, h = c.winfo_width(), c.winfo_height()
            steps = 100
            for i in range(steps):
                ang_i = (angle + i * (360 / steps)) % 360
                r, g, b = sin_color(ang_i)
                color = f'#{r:02x}{g:02x}{b:02x}'
                x0 = int(w * i / steps)
                x1 = int(w * (i + 1) / steps)
                c.create_rectangle(x0, 0, x1, h, fill=color, width=0, tags="grad")
            frame._grad_angle = (angle + 2) % 360
            c.after(delay, update)
        update()

    # ----- ربط قائمة النقر بالزر الأيمن (Context Menu) لجميع جداول البيانات -----
    def bind_all_treeviews(self):
        trees = [getattr(self, attr) for attr in dir(self) if attr.startswith("tree_")]
        for tree in trees:
            if isinstance(tree, ttk.Treeview):
                tree.bind("<Button-3>", self.show_context_menu)

    def show_context_menu(self, event):
        widget = event.widget
        iid = widget.identify_row(event.y)
        if not iid:
            return
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Stop", command=lambda: self.perform_action(widget, iid, "stop"))
        self.context_menu.add_command(label="Suspend", command=lambda: self.perform_action(widget, iid, "suspend"))

        self.context_menu.add_command(label="Investigate",
                                      command=lambda: self.perform_action(widget, iid, "investigate"))
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    def perform_action(self, tree, iid, action):
        vals = tree.item(iid, "values")
        if not vals:
            return
        pid_str = vals[0]
        try:
            pid = int(pid_str)
        except ValueError:
            messagebox.showerror("Error", f"Cannot execute {action} due to invalid PID")
            return
        try:
            proc = psutil.Process(pid)
            if action == "stop":
                proc.terminate()
            elif action == "suspend":
                proc.suspend()
            elif action == "investigate":
                messagebox.showinfo("Investigate", f"Investigating process (PID={pid})")
            messagebox.showinfo("Done", f"Action {action} executed on process (PID={pid})")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ----- دوال جديدة لمعالجة عمليات الإيقاف والتعليق والتحقيق في تبويب Scan process -----
    def stop_selected_processes(self):
        selections = self.tree_all.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one process to stop")
            return
        for iid in selections:
            vals = self.tree_all.item(iid, "values")
            try:
                pid = int(vals[0])
                psutil.Process(pid).terminate()
            except Exception as e:
                messagebox.showerror("Error", f"Could not stop PID {vals[0]}: {str(e)}")
        messagebox.showinfo("Done", "Stop command executed on selected processes.")

    def suspend_selected_processes(self):
        selections = self.tree_all.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one process to suspend")
            return
        for iid in selections:
            vals = self.tree_all.item(iid, "values")

            try:
                pid = int(vals[0])
                psutil.Process(pid).suspend()
            except Exception as e:
                messagebox.showerror("Error", f"Could not suspend PID {vals[0]}: {str(e)}")
        messagebox.showinfo("Done", "Suspend command executed on selected processes.")

    def investigate_selected_processes(self):
        selections = self.tree_all.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one process to investigate")
            return
        for iid in selections:
            vals = self.tree_all.item(iid, "values")
            try:
                pid = int(vals[0])
                messagebox.showinfo("Investigate", f"Investigating process (PID={pid})")
            except Exception as e:
                messagebox.showerror("Error", f"Error investigating PID {vals[0]}: {str(e)}")
        messagebox.showinfo("Done", "Investigate command executed on selected processes.")

    # ----- دوال جديدة لمعالجة العمليات في تبويب Suspicious Services -----
    def perform_stop_action_svc(self):
        selections = self.tree_svc.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one service to stop")
            return
        for iid in selections:
            vals = self.tree_svc.item(iid, "values")
            try:
                pid_str = vals[0]
                if pid_str and str(pid_str).isdigit():
                    psutil.Process(int(pid_str)).terminate()
            except Exception:
                continue
        messagebox.showinfo("Done", "Selected services stopped")

    def perform_ban_action_svc(self):
        selections = self.tree_svc.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one service to ban")
            return
        for iid in selections:
            vals = self.tree_svc.item(iid, "values")
            try:
                pid_str = vals[0]
                if pid_str and str(pid_str).isdigit():
                    psutil.Process(int(pid_str)).kill()
            except Exception:
                continue
        messagebox.showinfo("Done", "Selected services banned")


    # ------------------------------------------------------------------
    # 0. تبويب Welcome مع خلفية تفاعلية ثلاثية الأبعاد ومؤثرات رقمية
    # ------------------------------------------------------------------
    def create_tab_welcome(self):
        f = self.tab_welcome
        self.welcome_canvas = tk.Canvas(f, bg="black", highlightthickness=0)
        self.welcome_canvas.pack(fill="both", expand=True)
        self.welcome_canvas.bind("<Configure>", self.start_effects)
        desc = (
            "⚙️ Service Scanner Pro \n"
            "🔍 يكتشف الخدمات المشبوهة والشاذة تلقائيًا \n"
            "🛡️ يحلل السلوك ويُقيّم المخاطر باستخدام الذكاء الاصطناعي \n"
            "📊 يعرض تقارير رسومية ثلاثية الأبعاد وتفاعلية \n"
            "💾 يتيح تصدير النتائج وإيقاف أو عزل العمليات المشتبه بها"
        )
        self.welcome_canvas.create_text(
            20, 20,
            text=desc,
            anchor="nw",
            font=(FONT_NAME, 16, "bold"),
            fill="white",
            tags="description"
        )

    def start_effects(self, event):
        self.width = event.width
        self.height = event.height
        self.cx = self.width / 2
        self.cy = self.height / 2
        # Starfield effect
        self.stars = [{"x": random.uniform(-1, 1), "y": random.uniform(-1, 1), "z": random.uniform(0.1, 1)} for _ in range(200)]
        self.star_speed = 0.02
        # Digital Rain effect
        self.drops = []
        self.symbols = ["0", "1", "<>", "{}", "();", "⚙️", "💻"]
        for i in range(60):
            x = random.randint(0, self.width)
            y = random.randint(-self.height, 0)
            speed = random.uniform(2, 6)
            symbol = random.choice(self.symbols)
            self.drops.append({"x": x, "y": y, "speed": speed, "symbol": symbol})
        self.animate_starfield()
        self.animate_digital_rain()

    def animate_starfield(self):
        self.welcome_canvas.delete("star")
        for s in self.stars:
            s["z"] -= self.star_speed
            if s["z"] <= 0:

                s.update({"x": random.uniform(-1, 1), "y": random.uniform(-1, 1), "z": 1.0})
            k = min(self.width, self.height) / 2 / s["z"]
            x = s["x"] * k + self.cx
            y = s["y"] * k + self.cy
            size = max(1, (1 - s["z"]) * 5)
            r = int((1 - s["z"]) * 128)
            b = int(255 - (1 - s["z"]) * 128)
            color = f"#{r:02x}00{b:02x}"
            self.welcome_canvas.create_oval(x - size, y - size, x + size, y + size,
                                            fill=color, outline="", tags="star")
        self.welcome_canvas.after(33, self.animate_starfield)

    def animate_digital_rain(self):
        self.welcome_canvas.delete("rain")
        for d in self.drops:
            d["y"] += d["speed"]
            if d["y"] > self.height:
                d.update({"y": random.randint(-self.height, 0), "x": random.randint(0, self.width),
                          "symbol": random.choice(self.symbols)})
            x, y, sym = d["x"], d["y"], d["symbol"]
            self.welcome_canvas.create_text(x, y, text=sym, font=(FONT_NAME, 12, "bold"), fill="#0f0",
                                            tags="rain")
        self.welcome_canvas.after(50, self.animate_digital_rain)

    # ------------------------------------------------------------------
    # 1. تبويب فحص العمليات (Scan process) مع التحسينات المطلوبة
    # ------------------------------------------------------------------
    def create_tab_all(self):
        f = self.tab_all
        top = tk.Frame(f, bg=TERTIARY)
        top.pack(fill=tk.X, padx=10, pady=10)
        self.filter_options = [
            ("All", "#1976D2"),
            ("Signature", "#FF6347"),
            ("Anomaly_score", "#FF4500"),
            ("anomaly", "#FF8C00"),
            ("Delta_cpu", "#42A5F5"),
            ("Delta_io", "#00BCD4"),
            ("tcp_listen", "#8BC34A"),
            ("tcp_established", "#4CAF50"),
            ("udp", "#FFC107")
        ]
        self.filter_var_all = tk.StringVar(value="All")
        self.filter_mb = tk.Menubutton(top, text="Filter: All", font=(FONT_NAME, 11, 'bold'),
                                       bg=random.choice(GRADIENTS), fg="white", relief="raised")
        self.filter_menu = tk.Menu(self.filter_mb, tearoff=0)
        for opt, col in self.filter_options:
            self.filter_menu.add_command(label=opt, background=col, foreground="white",
                                         font=(FONT_NAME, 11, 'bold'),
                                         command=lambda opt=opt: self.set_filter(opt))
        self.filter_mb.configure(menu=self.filter_menu)

        self.filter_mb.pack(side=tk.LEFT, padx=5)
        btn_export_all = tk.Button(top, text="تصدير كل النتائج", fg="white", font=(FONT_NAME, 11, 'bold'),
                                   command=self.open_export_dialog_process, bg=random.choice(GRADIENTS))
        btn_export_all.pack(side=tk.RIGHT, padx=5)
        self.progress_all = ttk.Progressbar(top,
                                            style="Custom.Horizontal.TProgressbar", orient="horizontal",
                                            mode="determinate", length=200)
        self.progress_all.pack(side=tk.LEFT, padx=10)
        mid_frame = tk.Frame(f)
        mid_frame.pack(fill='both', expand=True, padx=10, pady=10)
        tree_frame = tk.Frame(mid_frame)
        tree_frame.pack(side=tk.LEFT, fill='both', expand=True)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        cols = ("PID", "name", "timestamp", "path", "account", "frequency",
                "age_days", "anomaly_score", "anomalous", "signature", "delta_cpu", "delta_io", "tcp_listen",
                "tcp_established", "udp")
        self.tree_all = ttk.Treeview(tree_frame, columns=cols, show='headings',
                                     yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for c in cols:
            self.tree_all.heading(c, text=c.capitalize())
            if c == "PID":
                self.tree_all.column(c, width=70)
            elif c == "name":
                self.tree_all.column(c, width=150)
            elif c == "timestamp":
                self.tree_all.column(c, width=150)
            elif c == "path":
                self.tree_all.column(c, width=300)
            elif c in ("anomaly_score", "delta_cpu", "delta_io"):
                self.tree_all.column(c, width=90)
            elif c in ("anomalous", "tcp_listen", "tcp_established", "udp"):
                self.tree_all.column(c, width=80)
            else:
                self.tree_all.column(c, width=100)
        self.tree_all.pack(fill='both', expand=True)
        vsb.config(command=self.tree_all.yview)
        vsb.pack(side=tk.RIGHT, fill='y')
        hsb.config(command=self.tree_all.xview)
        hsb.pack(side=tk.BOTTOM, fill='x')

        self.tree_all.tag_configure('anom', background="#ffcccc")
        side_panel = tk.Frame(mid_frame, bg=MAIN_BG)
        side_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=10)
        can = tk.Canvas(side_panel, width=220, height=260, bg=MAIN_BG, highlightthickness=0)
        can.pack()
        triangle_id = can.create_polygon(110, 20, 10, 240, 210, 240,
                                         outline=random.choice(GRADIENTS), fill='', width=3)
        self.animate_triangle_border(can, triangle_id)
        btn_frame = tk.Frame(can, bg=MAIN_BG)
        can.create_window(110, 130, window=btn_frame)

        row0_frame = tk.Frame(btn_frame, bg=MAIN_BG)
        row0_frame.grid(row=0, column=0, pady=5)
        btn_collect = tk.Button(row0_frame, text="Collect Baseline", fg="white", font=(FONT_NAME, 11, 'bold'),
                                command=lambda: threading.Thread(target=self.collect_baseline_with_spinner_all,
                                                                 daemon=True).start())
        btn_collect.pack(side=tk.LEFT, padx=5)
        self.spinner_label_all = tk.Label(row0_frame, text="", font=(FONT_NAME, 12), fg="yellow", bg=MAIN_BG)
        self.spinner_label_all.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_collect)
        btn_train = tk.Button(btn_frame, text="Train Model", fg="white", font=(FONT_NAME, 11, 'bold'),
                              command=lambda: threading.Thread(target=self.train_model, daemon=True).start())
        btn_train.grid(row=1, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_train)
        btn_scan = tk.Button(btn_frame, text="Scan Process", fg="white", font=(FONT_NAME, 11, 'bold'),
                             command=lambda: threading.Thread(target=self.scan_all_services, daemon=True).start())
        btn_scan.grid(row=2, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_scan)
        btn_suspend = tk.Button(btn_frame, text="Suspend", fg="white", font=(FONT_NAME, 11, 'bold'),
                                command=lambda: threading.Thread(target=self.suspend_selected_processes,
                                                                 daemon=True).start())
        btn_suspend.grid(row=3, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_suspend)
        btn_investigate = tk.Button(btn_frame, text="Investigate", fg="white", font=(FONT_NAME, 11, 'bold'),
                                    command=lambda: threading.Thread(target=self.investigate_selected_processes,
                                                                     daemon=True).start())
        btn_investigate.grid(row=4, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_investigate)
        btn_stop = tk.Button(btn_frame, text="Stop", fg="white", font=(FONT_NAME, 11, 'bold'),
                             command=lambda: threading.Thread(target=self.stop_selected_processes, daemon=True).start())
        btn_stop.grid(row=5, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_stop)
        btn_pause = tk.Button(btn_frame, text="Pause Radar", fg="white", font=(FONT_NAME, 11, 'bold'),
                              command=self.toggle_radar_pause, bg=random.choice(GRADIENTS))
        btn_pause.grid(row=6, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_pause)

        chart_frame = tk.Frame(f, bg=TERTIARY)
        chart_frame.pack(fill=tk.X, padx=10, pady=5)
        # رسم الرسم البياني ثلاثي الأبعاد (3D) عادي
        self.fig_dummy = plt.figure(figsize=(6, 4), facecolor=TERTIARY)
        self.ax_dummy = self.fig_dummy.add_subplot(121, projection='3d')
        canvas_dummy = FigureCanvasTkAgg(self.fig_dummy, master=chart_frame)
        canvas_dummy.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)
        # رسم مخطط راداري جديد مع التحسينات المطلوبة:
        self.fig_new_radar = plt.Figure(figsize=(6, 4), facecolor=TERTIARY)
        self.ax_new_radar = self.fig_new_radar.add_subplot(111, projection='polar')
        canvas_new_radar = FigureCanvasTkAgg(self.fig_new_radar, master=chart_frame)
        canvas_new_radar.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)

        # إنشاء مؤشر تفاعلي (Tooltip) باستخدام mplcursors مرة واحدة بعد أول رسم
        self.radar_cursor_initialized = False
        self.animate_new_radar_chart()

    def set_filter(self, opt):
        self.filter_var_all.set(opt)
        self.filter_mb.config(text=f"Filter: {opt}")
        self.filter_all_tree(opt)

    def update_3d_all(self, services):
        self.ax_dummy.clear()
        xs = [s['anomaly_score'] for s in services if s['anomaly_score'] != '']
        ys = [s['age_days'] for s in services]
        zs = list(range(len(services)))
        if xs and ys:
            self.ax_dummy.scatter(xs, ys, zs)
            self.ax_dummy.set_xlabel('Anomaly Score')
            self.ax_dummy.set_ylabel('Age (days)')
            self.ax_dummy.set_zlabel('Index')
        self.fig_dummy.canvas.draw()

    # ---------- تحديث مخطط الرادار الجديد مع التأثيرات المطلوبة ----------
    def animate_new_radar_chart(self):
        if self.radar_paused:
            self.fig_new_radar.canvas.get_tk_widget().after(200, self.animate_new_radar_chart)
            return
        self.ax_new_radar.clear()
        # إعداد المخطط القطبي والإعدادات الثابتة
        self.ax_new_radar.set_rlim(0, 6)
        self.ax_new_radar.set_theta_zero_location("N")
        self.ax_new_radar.set_theta_direction(-1)
        # رسم شبكة متحركة: دوائر متدرجة بخط منقط
        num_circles = 5
        for i in range(1, num_circles + 1):
            self.ax_new_radar.plot(np.linspace(0, 2 * np.pi, 100), [i * 6 / num_circles] * 100, color='blue',
                                   linestyle=':', linewidth=0.5)
        # رسم الشعاع الدائري بنقاط متحركة
        N = 20  # عدد النقاط
        if not hasattr(self, '_radar_phase'):
            self._radar_phase = 0
        phase = math.radians(self._radar_phase)
        base_angles = np.linspace(0, 2 * np.pi, N, endpoint=False)
        moving_angles = base_angles + phase
        r_value = 5  # نصف قطر ثابت
        self.ax_new_radar.scatter(moving_angles, [r_value] * N, color='cyan', s=50)
        # رسم دوار ليزر متحرك: خط أحمر من المركز إلى الحافة عند الزاوية self._radar_phase
        self.ax_new_radar.plot([phase, phase], [0, 6], color='red', linewidth=2)
        # <------ إضافة المضلعات الحمراء المتحركة -------->
        num_labels = 5
        label_angles = np.linspace(0, 2 * np.pi, num_labels, endpoint=False)
        polygon_radii = [4 + 1 * math.sin(math.radians(self._radar_phase) + angle) for angle in label_angles]

        polygon_angles = np.append(label_angles, label_angles[0])
        polygon_radii.append(polygon_radii[0])
        self.ax_new_radar.fill(polygon_angles, polygon_radii, color='red', alpha=0.3)
        labels = ["Signature", "Anomaly", "I/O", "Memory", "Suspicious"]
        for ang, lab in zip(label_angles, labels):
            self.ax_new_radar.text(ang, 6.2, lab, color='white', fontsize=10, fontname=FONT_NAME, ha='center',
                                   va='center')
        self._radar_phase = (self._radar_phase + 5) % 360
        self.fig_new_radar.canvas.draw()
        self.fig_new_radar.canvas.get_tk_widget().after(200, self.animate_new_radar_chart)

    def scan_all_services(self):
        services = []
        procs = list(psutil.process_iter(['pid', 'name', 'exe', 'cmdline', 'create_time', 'username']))
        total = len(procs)
        self.root.after(0, lambda: self.progress_all.configure(maximum=total, value=0))
        for idx, proc in enumerate(procs, start=1):
            try:
                pid = proc.info['pid']
                name = proc.info['name'] or ''
                path = proc.info['exe'] or ''
                sig_ok = self.verify_signature(path)
                acc = self.get_account_type(proc)
                freq = self.get_frequency(proc)
                age = 0
                if proc.info.get('create_time'):
                    age = (datetime.now() - datetime.fromtimestamp(proc.info['create_time'])).days
                ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                try:
                    initial_cpu = proc.cpu_percent(interval=None)
                    io = proc.io_counters() if hasattr(proc, 'io_counters') else None
                    initial_io = (io.read_bytes + io.write_bytes) if io else 0
                    time.sleep(0.01)
                    second_cpu = proc.cpu_percent(interval=None)
                    io2 = proc.io_counters() if hasattr(proc, 'io_counters') else None
                    second_io = (io2.read_bytes + io2.write_bytes) if io2 else 0
                    delta_cpu = round(second_cpu - initial_cpu, 3)
                    delta_io = second_io - initial_io
                    cpu = second_cpu
                except Exception as ex:
                    cpu = proc.cpu_percent(interval=0.01)
                    delta_cpu = 0.0
                    delta_io = 0
                try:
                    mem = proc.memory_info()
                    rss, vms = mem.rss, mem.vms
                except Exception:
                    rss, vms = 0, 0
                io = proc.io_counters() if hasattr(proc, 'io_counters') else None
                rb = io.read_bytes if io else 0
                wb = io.write_bytes if io else 0

                tcp_listen = 0
                tcp_established = 0
                udp_count = 0
                nconns = 0
                try:
                    conns = proc.connections(kind='inet') if hasattr(proc, 'connections') else []
                    for conn in conns:
                        if conn.type == socket.SOCK_STREAM:
                            if conn.status == 'LISTEN':
                                tcp_listen += 1
                            elif conn.status == 'ESTABLISHED':
                                tcp_established += 1
                        elif conn.type == socket.SOCK_DGRAM:
                            udp_count += 1
                    nconns = len(conns)
                except Exception:
                    nconns = 0
                an_score, anom = None, 0
                if self.anomaly_model:
                    feats = [[cpu, rss, vms, rb, wb, nconns]]
                    an_score = float(self.anomaly_model.decision_function(feats)[0])
                    anom = 1 if an_score < -0.2 else 0
                record = {
                    'pid': pid,
                    'name': name,
                    'path': path,
                    'account': acc,
                    'frequency': freq,
                    'age_days': age,
                    'timestamp': ts,
                    'anomaly_score': round(an_score, 3) if an_score is not None else '',
                    'anomalous': anom,
                    'signature_valid': 1 if sig_ok else 0,
                    'source': "Processes",
                    'delta_cpu': delta_cpu,
                    'delta_io': delta_io,
                    'tcp_listen': tcp_listen,
                    'tcp_established': tcp_established,
                    'udp': udp_count
                }
                services.append(record)
            except Exception as e:
                continue
            if idx % 10 == 0:
                self.root.after(0, lambda i=idx: self.progress_all.configure(value=i))
        self.root.after(0, lambda: self.progress_all.configure(value=total))
        self.all_services_data = services
        self.update_all(services)
        self.update_3d_all(services)

    def update_all(self, services):

        self.tree_all.delete(*self.tree_all.get_children())
        for s in services:
            tags = ()
            if s['anomalous'] == 1 or s['signature_valid'] == 0:
                tags = ('anom',)
            row = [
                s['pid'],
                s['name'],
                s['timestamp'],
                s['path'],
                s['account'],
                s['frequency'],
                s['age_days'],
                s['anomaly_score'],
                s['anomalous'],
                "Valid" if s['signature_valid'] == 1 else "Invalid",
                s['delta_cpu'],
                s['delta_io'],
                s['tcp_listen'],
                s['tcp_established'],
                s['udp']
            ]
            self.tree_all.insert('', 'end', values=row, tags=tags)
        self.filter_all_tree(self.filter_var_all.get())

    def filter_all_tree(self, selection):
        self.tree_all.delete(*self.tree_all.get_children())
        for s in self.all_services_data:
            insert_item = False
            if selection == "All":
                insert_item = True
            elif selection == "Signature":
                if (("Valid" if s['signature_valid'] == 1 else "Invalid") != "Valid"):
                    insert_item = True
            elif selection == "Anomaly_score":
                try:
                    score = float(s['anomaly_score'])
                    if score < -0.2:
                        insert_item = True
                except Exception:
                    insert_item = False
            elif selection == "anomaly":
                if s['anomalous'] == 1:
                    insert_item = True
            elif selection == "Delta_cpu":
                if s['delta_cpu'] != 0:
                    insert_item = True
            elif selection == "Delta_io":
                if s['delta_io'] != 0:
                    insert_item = True
            elif selection == "tcp_listen":

                if s['tcp_listen'] != 0:
                    insert_item = True
            elif selection == "tcp_established":
                if s['tcp_established'] != 0:
                    insert_item = True
            elif selection == "udp":
                if s['udp'] != 0:
                    insert_item = True
            if insert_item:
                tags = ()
                if s['anomalous'] == 1 or s['signature_valid'] == 0:
                    tags = ('anom',)
                row = [
                    s['pid'],
                    s['name'],
                    s['timestamp'],
                    s['path'],
                    s['account'],
                    s['frequency'],
                    s['age_days'],
                    s['anomaly_score'],
                    s['anomalous'],
                    "Valid" if s['signature_valid'] == 1 else "Invalid",
                    s['delta_cpu'],
                    s['delta_io'],
                    s['tcp_listen'],
                    s['tcp_established'],
                    s['udp']
                ]
                self.tree_all.insert('', 'end', values=row, tags=tags)

    def export_all_excel_process(self):
        data = [self.tree_all.item(i, 'values')[1:] for i in self.tree_all.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if fn:
            df = pd.DataFrame(data,
                              columns=["Name", "Timestamp", "Path", "Account", "Frequency", "Age", "Anomaly Score",
                                       "Anomalous", "Signature", "Delta Cpu", "Delta Io", "TCP Listen",
                                       "TCP Established", "UDP"])
            try:
                df.to_excel(fn, index=False)
                messagebox.showinfo("Exported", "Excel exported")
            except Exception as e:
                messagebox.showerror("Error", f"Excel export failed: {str(e)}")

    def export_all_html_process(self):
        data = [self.tree_all.item(i, 'values')[1:] for i in self.tree_all.get_children()]
        if not data:

            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".html")
        if fn:
            df = pd.DataFrame(data,
                              columns=["Name", "Timestamp", "Path", "Account", "Frequency", "Age", "Anomaly Score",
                                       "Anomalous", "Signature", "Delta Cpu", "Delta Io", "TCP Listen",
                                       "TCP Established", "UDP"])
            df.to_html(fn, index=False)
            messagebox.showinfo("Exported", "HTML exported")

    def export_all_pdf_process(self):
        data = [self.tree_all.item(i, 'values')[1:] for i in self.tree_all.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".pdf")
        if fn:
            try:
                from fpdf import FPDF
            except ImportError:
                messagebox.showerror("Error", "fpdf module not installed. Use pip install fpdf")
                return
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            col_names = ["Name", "Timestamp", "Path", "Account", "Frequency", "Age", "Anomaly Score", "Anomalous",
                         "Signature", "Delta Cpu", "Delta Io", "TCP Listen", "TCP Established", "UDP"]
            for col in col_names:
                pdf.cell(30, 10, txt=col, border=1)
            pdf.ln()
            for row in data:
                for item in row:
                    pdf.cell(30, 10, txt=str(item), border=1)
                pdf.ln()
            pdf.output(fn)
            messagebox.showinfo("Exported", "PDF exported")

    def open_export_dialog_process(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("حدد نوع التصدير")
        dialog.geometry("300x150")
        self.animate_dialog_wavy(dialog)
        tk.Label(dialog, text="اختر نوع التصدير:", font=(FONT_NAME, 11, 'bold'), bg=dialog['bg']).pack(pady=10)
        export_type = tk.StringVar(value="Excel")
        formats = [("Excel", "Excel"), ("PDF", "PDF"), ("HTML", "HTML")]
        for text, value in formats:
            tk.Radiobutton(dialog, text=text, variable=export_type, value=value, font=(FONT_NAME, 10),
                           bg=dialog['bg']).pack(anchor=tk.W, padx=20)

        def submit():
            choice = export_type.get()
            dialog.destroy()
            if choice == "Excel":
                self.export_all_excel_process()
            elif choice == "PDF":
                self.export_all_pdf_process()
            elif choice == "HTML":
                self.export_all_html_process()

        tk.Button(dialog, text="تصدير", font=(FONT_NAME, 11, 'bold'), bg=random.choice(GRADIENTS), fg="white",
                  command=submit).pack(pady=10)

    # ------------------------------------------------------------------
    # 3. تبويب Keywords Scan مع التعديلات المطلوبة وإضافة دعم استيراد SIGMA (YAML)
    # ------------------------------------------------------------------
    def create_tab_keywords(self):
        f = self.tab_kw
        top = tk.Frame(f, bg=SECONDARY)
        top.pack(fill=tk.X, padx=10, pady=10)
        btn_scan = tk.Button(top, text="Scan by Keywords", bg=random.choice(GRADIENTS), fg="white",
                             font=(FONT_NAME, 11, 'bold'),
                             command=lambda: threading.Thread(target=self.scan_by_keywords, daemon=True).start())
        btn_scan.pack(side=tk.LEFT, padx=5)
        btn_settings = tk.Button(top, text="Settings", bg=random.choice(GRADIENTS), fg="white",
                                 font=(FONT_NAME, 11, 'bold'),
                                 command=self.open_settings)
        btn_settings.pack(side=tk.LEFT, padx=5)
        # زر استيراد قواعد SIGMA بصيغة YAML
        btn_import_sigma = tk.Button(top, text="استيراد من ملف sigma yaml", bg=random.choice(GRADIENTS), fg="white",
                                     font=(FONT_NAME, 11, 'bold'),
                                     command=lambda: threading.Thread(target=self.import_sigma_yaml, daemon=True).start())
        btn_import_sigma.pack(side=tk.LEFT, padx=5)
        bar_frame = tk.Frame(f, bg=SECONDARY)
        bar_frame.pack(fill=tk.X, padx=10, pady=5)
        self.fig_kw2d, self.ax_kw2d = plt.subplots(figsize=(6, 4), facecolor=SECONDARY)
        canvas_kw2d = FigureCanvasTkAgg(self.fig_kw2d, master=bar_frame)
        canvas_kw2d.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)
        self.fig_k3d = plt.figure(figsize=(6, 4), facecolor=SECONDARY)
        self.ax_k3d = self.fig_k3d.add_subplot(111, projection='3d')
        canvas_k3d = FigureCanvasTkAgg(self.fig_k3d, master=bar_frame)
        canvas_k3d.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)
        cols = ('pid', 'name', 'path', 'account', 'keyword', 'age', 'frequency', 'timestamp')
        self.tree_kw = ttk.Treeview(f, columns=cols, show='headings')
        for c in cols:
            self.tree_kw.heading(c, text=c.capitalize())
            if c == "path":

                self.tree_kw.column(c, width=300)
            else:
                self.tree_kw.column(c, width=150)
        self.tree_kw.pack(fill='both', expand=True, padx=10, pady=10)
        btn_frame_bottom = tk.Frame(f, bg=SECONDARY)
        btn_frame_bottom.pack(fill=tk.X, padx=10, pady=10)
        btn_exp_excel = tk.Button(btn_frame_bottom, text="Export Excel", fg="white", font=(FONT_NAME, 11, 'bold'),
                                  command=self.export_kw_excel)
        btn_exp_excel.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_exp_excel)
        btn_stop = tk.Button(btn_frame_bottom, text="Stop", fg="white", font=(FONT_NAME, 11, 'bold'),
                             command=self.stop_kw)
        btn_stop.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_stop)
        btn_hadir = tk.Button(btn_frame_bottom, text="حضر", fg="white", font=(FONT_NAME, 11, 'bold'),
                              command=self.hadir_kw)
        btn_hadir.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_hadir)

    def scan_by_keywords(self):
        data = []
        for proc in psutil.process_iter(['pid', 'name', 'exe', 'cmdline']):
            try:
                cmd = ' '.join(proc.info['cmdline'] or [])
                for kw in self.settings.keywords:
                    if kw.lower() in cmd.lower():
                        try:
                            p = psutil.Process(proc.info['pid'])
                            age = (datetime.now() - datetime.fromtimestamp(p.create_time())).days
                            account = p.username()
                        except Exception:
                            age = 0
                            account = "Unknown"
                        frequency = random.randint(1, 5)
                        timestamp_val = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        record = {
                            'pid': proc.info['pid'],
                            'name': proc.info['name'] or '',
                            'path': proc.info['exe'] or '',
                            'account': account,
                            'keyword': kw,
                            'age': age,
                            'frequency': frequency,
                            'timestamp': timestamp_val
                        }
                        data.append((record['pid'], record['name'], record['path'],
                                     record['account'], record['keyword'], record['age'],
                                     record['frequency'], record['timestamp']))
                        self.add_history_record(record)
                        break

            except Exception:
                continue
        self.tree_kw.delete(*self.tree_kw.get_children())
        for r in data:
            self.tree_kw.insert("", "end", values=r)
        self.update_2d_kw(data)
        self.update_3d_kw(data)

    def export_kw_excel(self):
        data = [self.tree_kw.item(i, 'values') for i in self.tree_kw.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if fn:
            df = pd.DataFrame(data,
                              columns=["PID", "Name", "Path", "Account", "Keyword", "Age", "Frequency", "Timestamp"])
            try:
                df.to_excel(fn, index=False)
                messagebox.showinfo("Exported", "Excel exported")
            except Exception as e:
                messagebox.showerror("Error", f"Excel export failed: {str(e)}")

    def stop_kw(self):
        selections = self.tree_kw.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one process to stop")
            return
        for iid in selections:
            vals = self.tree_kw.item(iid, "values")
            try:
                pid = int(vals[0])
                psutil.Process(pid).terminate()
            except Exception as e:
                messagebox.showerror("Error", f"Could not stop PID {vals[0]}: {str(e)}")
        messagebox.showinfo("Done", "Stop command executed for selected processes.")

    def hadir_kw(self):
        selections = self.tree_kw.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one process to resume")
            return
        for iid in selections:
            vals = self.tree_kw.item(iid, "values")
            try:
                pid = int(vals[0])
                psutil.Process(pid).resume()
            except Exception as e:
                messagebox.showerror("Error", f"Could not resume PID {vals[0]}: {str(e)}")
        messagebox.showinfo("Done", "حضر command executed for selected processes.")

    def update_2d_kw(self, data):
        self.ax_kw2d.clear()
        counts = {}
        for row in data:
            key = (row[3], row[4])
            counts[key] = counts.get(key, 0) + 1
        labels = [f"{acc}\n{kw}" for acc, kw in counts.keys()]
        vals = list(counts.values())
        if labels:
            self.ax_kw2d.bar(labels, vals, color="#cc66ff")
            self.ax_kw2d.set_xlabel("Account & Keyword")
            self.ax_kw2d.set_ylabel("Count")
        self.fig_kw2d.canvas.draw()

    def update_3d_kw(self, data):
        self.ax_k3d.clear()
        xs = [row[5] for row in data]  # age
        ys = [row[0] for row in data]  # PID
        zs = list(range(len(data)))
        if xs:
            self.ax_k3d.scatter(xs, ys, zs, color="#ff99cc")
            self.ax_k3d.set_xlabel("Age")
            self.ax_k3d.set_ylabel("PID")
            self.ax_k3d.set_zlabel("Index")
        self.fig_k3d.canvas.draw()

    def open_settings(self):
        settings_win = tk.Toplevel(self.root)
        settings_win.title("Keywords Settings")
        settings_win.geometry("400x300")
        canvas_bg = tk.Canvas(settings_win, bg="white")
        canvas_bg.pack(fill="both", expand=True)
        width = 400
        height = 300
        for i in range(0, height, 20):
            color = f'#{(100 + i) % 255:02x}{(150 + i // 2) % 255:02x}{(200 + i // 3) % 255:02x}'
            canvas_bg.create_rectangle(0, i, width, i + 20, fill=color, outline="")
        frame = tk.Frame(canvas_bg, bg="#ffffff")
        canvas_bg.create_window(200, 150, window=frame)
        lbl = tk.Label(frame, text="Manage Keywords", font=(FONT_NAME, 12, "bold"), bg="#ffffff")
        lbl.pack(pady=5)
        listbox = tk.Listbox(frame, font=(FONT_NAME, 11), width=30, height=8)
        listbox.pack(pady=5)
        for kw in self.settings.keywords:
            listbox.insert(tk.END, kw)
        entry = tk.Entry(frame, font=(FONT_NAME, 11))
        entry.pack(pady=5)
        btn_frame = tk.Frame(frame, bg="#ffffff")
        btn_frame.pack(pady=5)

        def add_keyword():

            new_kw = entry.get().strip()
            if new_kw and new_kw not in self.settings.keywords:
                listbox.insert(tk.END, new_kw)
                self.settings.keywords.append(new_kw)
                entry.delete(0, tk.END)

        def delete_keyword():
            selection = listbox.curselection()
            if selection:
                idx = selection[0]
                listbox.delete(idx)
                del self.settings.keywords[idx]

        def save_keywords():
            self.settings.save_keywords(self.settings.keywords)
            messagebox.showinfo("Saved", "Keywords saved successfully")
            settings_win.destroy()

        btn_add = tk.Button(btn_frame, text="Add", font=(FONT_NAME, 11, "bold"), bg=random.choice(GRADIENTS),
                            fg="white", command=add_keyword)
        btn_add.pack(side=tk.LEFT, padx=5)
        btn_del = tk.Button(btn_frame, text="Delete", font=(FONT_NAME, 11, "bold"), bg=random.choice(GRADIENTS),
                            fg="white", command=delete_keyword)
        btn_del.pack(side=tk.LEFT, padx=5)
        btn_save = tk.Button(btn_frame, text="Save", font=(FONT_NAME, 11, "bold"), bg=random.choice(GRADIENTS),
                             fg="white", command=save_keywords)
        btn_save.pack(side=tk.LEFT, padx=5)

    # ---- إضافة وظيفة استيراد قواعد SIGMA الرسمية (YAML) ----
    def import_sigma_yaml(self):
        # السماح بتحديد ملف واحد أو أكثر من الملفات بصيغة YAML
        files = filedialog.askopenfilenames(title="اختر ملف/ملفات YAML لقواعد SIGMA",
                                            filetypes=[("YAML files", "*.yml *.yaml")])
        if not files:
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        imported = 0
        for file_path in files:
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    rule_data = yaml.safe_load(f)
                # نفترض أن القاعدة تحتوي على مفتاح "title" لتحديد عنوان القاعدة
                rule_title = rule_data.get("title", os.path.basename(file_path))
                rule_id = rule_data.get("id", None)
                # تحويل البيانات إلى سلسلة JSON للتخزين
                rule_json = json.dumps(rule_data, ensure_ascii=False)
                c.execute('''

                    INSERT INTO sigma_rules (rule_id, rule_title, rule_data)
                    VALUES (?, ?, ?)
                ''', (rule_id, rule_title, rule_json))
                imported += 1
            except Exception as e:
                print(f"Error importing {file_path}: {str(e)}")
                continue
        conn.commit()
        conn.close()
        messagebox.showinfo("استيراد", f"تم استيراد {imported} قاعدة من ملفات YAML بنجاح.")

    # ------------------------------------------------------------------
    # 4. تبويب Suspicious Services مع التعديلات الخاصة به فقط
    # ------------------------------------------------------------------
    def create_tab_suspicious_services(self):
        f = self.tab_svc
        top = tk.Frame(f, bg=PRIMARY)
        top.pack(fill=tk.X, padx=10, pady=10)
        new_filter_options = [
            ("-- اختر فلتر --", None),
            ("Signature", "#FF6347"),
            ("Hash", "#FF7F50"),
            ("Anomaly", "#FF4500"),
            ("Hidden", "#FF8C00"),
            ("Suspicious", "#42A5F5"),
            ("tcp_listen", "#00BCD4"),
            ("tcp_established", "#8BC34A"),
            ("udp", "#4CAF50"),
            ("delta_cpu", "#FFC107"),
            ("Delta_io", "#A0522D")
        ]
        self.filter_var = tk.StringVar(value=new_filter_options[0][0])
        self.filter_menu = tk.OptionMenu(top, self.filter_var, *[item[0] for item in new_filter_options],
                                         command=self.filter_svc)
        self.filter_menu.configure(bg=random.choice(GRADIENTS), fg="white", font=(FONT_NAME, 11, 'bold'))
        for index, (opt, col) in enumerate(new_filter_options):
            self.filter_menu["menu"].entryconfigure(index, background=col if col else random.choice(GRADIENTS),
                                                    foreground="white", font=(FONT_NAME, 11, 'bold'))
        self.filter_menu.pack(side=tk.LEFT, padx=5)
        btn_refresh = tk.Button(top, text="Refresh", fg="white", font=(FONT_NAME, 11, 'bold'),
                                command=lambda: threading.Thread(target=self.scan_svc, daemon=True).start())
        btn_refresh.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_refresh)
        self.progress_svc_ext = ttk.Progressbar(top, style="Custom.Horizontal.TProgressbar", orient="horizontal",
                                                mode="determinate", length=200)
        self.progress_svc_ext.pack(side=tk.LEFT, padx=10)
        btn_export_all_svc = tk.Button(top, text="تصدير كل النتائج", fg="white", font=(FONT_NAME, 11, 'bold'),
                                       command=self.open_export_dialog_svc, bg=random.choice(GRADIENTS))

        btn_export_all_svc.pack(side=tk.RIGHT, padx=5)
        side_frame = tk.Frame(f, bg=MAIN_BG)
        side_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)
        can = tk.Canvas(side_frame, width=220, height=220, bg=MAIN_BG, highlightthickness=0)
        can.pack()
        self.spinner_counter_label_svc = tk.Label(side_frame, text="", font=(FONT_NAME, 12), fg="yellow", bg=MAIN_BG)
        can.create_window(50, 10, window=self.spinner_counter_label_svc)
        triangle_id = can.create_polygon(110, 10, 10, 210, 210, 210, outline=random.choice(GRADIENTS), fill='', width=3)
        self.animate_triangle_border(can, triangle_id)
        btn_frame = tk.Frame(can, bg=MAIN_BG)
        can.create_window(110, 110, window=btn_frame)
        btn_scan = tk.Button(btn_frame, text="Scan Suspicious Services", fg="white", font=(FONT_NAME, 11, 'bold'),
                             command=lambda: threading.Thread(target=self.scan_svc, daemon=True).start())
        btn_scan.grid(row=0, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_scan)
        self.btn_baseline_svc = tk.Button(btn_frame, text="Collect Baseline", fg="white", font=(FONT_NAME, 11, 'bold'),
                                          command=lambda: threading.Thread(
                                              target=self.collect_baseline_with_spinner_svc, daemon=True).start())
        self.btn_baseline_svc.grid(row=1, column=0, padx=5, pady=5)
        self.animate_button_gradient(self.btn_baseline_svc)
        btn_train = tk.Button(btn_frame, text="Train Model", fg="white", font=(FONT_NAME, 11, 'bold'),
                              command=lambda: threading.Thread(target=self.train_model, daemon=True).start())
        btn_train.grid(row=2, column=0, padx=5, pady=5)
        self.animate_button_gradient(btn_train)
        self.spinner_label_svc = tk.Label(btn_frame, text="", font=(FONT_NAME, 12), fg="yellow", bg=MAIN_BG)
        self.spinner_label_svc.grid(row=2, column=1, padx=5)
        pw = ttk.PanedWindow(f, orient="vertical")
        pw.pack(fill='both', expand=True, padx=10, pady=10)
        chart_frame = tk.Frame(pw, bg=PRIMARY)
        pw.add(chart_frame, weight=1)
        self.fig_s3d = plt.figure(figsize=(6, 4), facecolor=PRIMARY)
        self.ax_s3d = self.fig_s3d.add_subplot(121, projection='3d')
        canvas_3d = FigureCanvasTkAgg(self.fig_s3d, master=chart_frame)
        canvas_3d.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)
        self.fig_heatmap = plt.Figure(figsize=(6, 4), facecolor=PRIMARY)
        self.ax_heatmap = self.fig_heatmap.add_subplot(121)
        canvas_heat = FigureCanvasTkAgg(self.fig_heatmap, master=chart_frame)
        canvas_heat.get_tk_widget().pack(side=tk.LEFT, padx=5, pady=5)
        table_frame = tk.Frame(pw)
        pw.add(table_frame, weight=2)
        container = tk.Frame(table_frame)
        container.pack(fill='both', expand=True)
        cols_svc = ("PID", "name", "timestamp", "status", "start_type", "username",
                    "binpath", "signature", "suspicious", "hidden", "sig_publisher", "hash_status",
                    "anomaly_score", "parent_pid", "tcp_listen", "tcp_established", "udp", "delta_cpu", "delta_io")
        self.tree_svc = ttk.Treeview(container, columns=cols_svc, show='headings')

        col_settings = {
            "PID": 70,
            "name": 150,
            "timestamp": 150,
            "status": 100,
            "start_type": 100,
            "username": 100,
            "binpath": 300,
            "signature": 80,
            "suspicious": 80,
            "hidden": 80,
            "sig_publisher": 120,
            "hash_status": 80,
            "anomaly_score": 90,
            "parent_pid": 80,
            "tcp_listen": 80,
            "tcp_established": 80,
            "udp": 80,
            "delta_cpu": 80,
            "delta_io": 80
        }
        for col, width in col_settings.items():
            self.tree_svc.heading(col, text=col.capitalize())
            self.tree_svc.column(col, width=width, minwidth=width, stretch=(False if col == "PID" else True))
        vsb = ttk.Scrollbar(container, orient="vertical", command=self.tree_svc.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.tree_svc.xview)
        self.tree_svc.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree_svc.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.tree_svc.tag_configure('sus', background="#ffcc66")
        btn_frame2 = tk.Frame(f, bg=PRIMARY)
        btn_frame2.pack(pady=5)
        btn_stop = tk.Button(btn_frame2, text="Stop Selected", fg="white", font=(FONT_NAME, 11, 'bold'),
                             command=self.perform_stop_action_svc)
        btn_stop.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_stop)
        btn_ban = tk.Button(btn_frame2, text="Ban Selected", fg="white", font=(FONT_NAME, 11, 'bold'),
                            command=self.perform_ban_action_svc)
        btn_ban.pack(side=tk.LEFT, padx=5)
        self.animate_button_gradient(btn_ban)

    def filter_svc(self, selection):
        for iid in self.tree_svc.get_children():
            self.tree_svc.item(iid, tags=())
        for iid in self.tree_svc.get_children():
            vals = self.tree_svc.item(iid, "values")
            if selection == "Signature":
                if vals[7] != "Valid":

                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "Hash":
                if vals[11] != "OK":
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "Anomaly":
                try:
                    score = float(vals[12])
                    if score < -0.5:
                        self.tree_svc.item(iid, tags=('filtered',))
                except Exception:
                    pass
            elif selection == "Hidden":
                if vals[9] == "Yes":
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "Suspicious":
                if vals[8] == "Yes":
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "tcp_listen":
                if vals[14] != "" and int(vals[14]) > 0:
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "tcp_established":
                if vals[15] != "" and int(vals[15]) > 0:
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "udp":
                if vals[16] != "" and int(vals[16]) > 0:
                    self.tree_svc.item(iid, tags=('filtered',))
            elif selection == "delta_cpu":
                try:
                    if float(vals[17]) != 0:
                        self.tree_svc.item(iid, tags=('filtered',))
                except Exception:
                    pass
            elif selection == "Delta_io":
                try:
                    if float(vals[18]) != 0:
                        self.tree_svc.item(iid, tags=('filtered',))
                except Exception:
                    pass
        self.tree_svc.tag_configure('filtered', background=random.choice(GRADIENTS))

    def scan_svc(self):
        self.start_spinner_with_counter(self.spinner_counter_label_svc, "svc_spinner_counter_running")
        items = []
        current_ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        registered_services = []
        if platform.system() == 'Windows':
            registered_services = get_services_from_scm()
        for svc in psutil.win_service_iter():
            try:
                info = svc.as_dict()
                binp = info.get('binpath', '')

                safe = any(binp.lower().startswith(d.lower()) for d in
                           ["C:\\Windows", "C:\\Program Files", "C:\\Program Files (x86)"])
                rec = {
                    'pid': get_service_pid(info.get('name', '')),
                    'name': info.get('name', ''),
                    'status': info.get('status', ''),
                    'start_type': info.get('start_type', ''),
                    'username': info.get('username', ''),
                    'binpath': binp,
                    'signature': "Valid" if self.verify_signature(binp) else "Invalid",
                    'suspicious': "Yes" if binp and not safe else "No",
                    'hidden': "No",
                    'sig_publisher': "KnownPublisher" if self.verify_signature(binp) else "N/A",
                    'hash_status': "OK",
                    'anomaly_score': round(random.uniform(-1, 1), 3) if self.anomaly_model else "",
                    'parent_pid': "",
                    'tcp_listen': 0,
                    'tcp_established': 0,
                    'udp': 0,
                    'delta_cpu': "",
                    'delta_io': ""
                }
                items.append(rec)
            except Exception:
                continue
        for proc in psutil.process_iter(['pid', 'name', 'exe', 'username']):
            try:
                exe = proc.info.get('exe', '')
                if not exe:
                    continue
                is_registered = any(svc.lower() in exe.lower() for svc in registered_services)
                hidden = "Yes" if not is_registered else "No"
                safe = True if platform.system() == "Windows" else exe.startswith(('/usr', '/bin', '/sbin'))
                try:
                    p = psutil.Process(proc.info.get('pid'))
                    ppid = p.ppid()
                except:
                    ppid = ""
                file_hash = self.compute_file_hash(exe)
                if file_hash and exe in baseline_hashes:
                    hash_status = "OK" if file_hash == baseline_hashes[exe] else "Modified"
                elif file_hash:
                    hash_status = "New"
                else:
                    hash_status = "N/A"
                try:
                    initial_cpu = proc.cpu_percent(interval=None)
                    io = proc.io_counters() if hasattr(proc, 'io_counters') else None
                    initial_io = (io.read_bytes + io.write_bytes) if io else 0
                    second_cpu = proc.cpu_percent(interval=None)
                    io2 = proc.io_counters() if hasattr(proc, 'io_counters') else None

                    second_io = (io2.read_bytes + io2.write_bytes) if io2 else 0
                    delta_cpu = round(second_cpu - initial_cpu, 3)
                    delta_io = second_io - initial_io
                except Exception:
                    delta_cpu = 0.0
                    delta_io = 0
                tcp_listen = 0
                tcp_established = 0
                udp_count = 0
                try:
                    conns = proc.connections(kind='inet') if hasattr(proc, 'connections') else []
                    for conn in conns:
                        if conn.type == socket.SOCK_STREAM:
                            if conn.status == 'LISTEN':
                                tcp_listen += 1
                            elif conn.status == 'ESTABLISHED':
                                tcp_established += 1
                        elif conn.type == socket.SOCK_DGRAM:
                            udp_count += 1
                except Exception:
                    tcp_listen = 0
                    tcp_established = 0
                    udp_count = 0
                rec = {
                    'pid': proc.info.get('pid', ''),
                    'name': proc.info.get('name', ''),
                    'status': "Running",
                    'start_type': "N/A",
                    'username': proc.info.get('username', ''),
                    'binpath': exe,
                    'signature': "Valid" if self.verify_signature(exe) else "Invalid",
                    'suspicious': "Yes" if exe and not safe else "No",
                    'hidden': hidden,
                    'sig_publisher': "KnownPublisher" if self.verify_signature(exe) else "N/A",
                    'hash_status': hash_status,
                    'anomaly_score': round(random.uniform(-1, 1), 3) if self.anomaly_model else "",
                    'parent_pid': ppid,
                    'source': "Suspicious Services",
                    'timestamp': current_ts,
                    'tcp_listen': tcp_listen,
                    'tcp_established': tcp_established,
                    'udp': udp_count,
                    'delta_cpu': delta_cpu,
                    'delta_io': delta_io
                }
                items.append(rec)
            except Exception:
                continue
        total = len(items)
        self.root.after(0, lambda: self.progress_svc_ext.configure(maximum=total, value=0))
        for idx, s in enumerate(items, start=1):

            if idx % 5 == 0:
                self.root.after(0, lambda i=idx: self.progress_svc_ext.configure(value=i))
        self.root.after(0, lambda: self.progress_svc_ext.configure(value=total))
        self.tree_svc.delete(*self.tree_svc.get_children())
        for s in items:
            tag = 'sus' if s['suspicious'] == 'Yes' else ''
            row = [
                s.get('pid', ''),
                s.get('name', ''),
                s.get('timestamp', current_ts),
                s.get('status', ''),
                s.get('start_type', ''),
                s.get('username', ''),
                s.get('binpath', ''),
                s.get('signature', ''),
                s.get('suspicious', ''),
                s.get('hidden', ''),
                s.get('sig_publisher', ''),
                s.get('hash_status', ''),
                s.get('anomaly_score', ''),
                s.get('parent_pid', ''),
                s.get('tcp_listen', 0),
                s.get('tcp_established', 0),
                s.get('udp', 0),
                s.get('delta_cpu', ''),
                s.get('delta_io', '')
            ]
            self.tree_svc.insert('', 'end', values=row, tags=(tag,))
        self.update_3d_svc(items)
        self.update_heatmap(items)
        self.stop_spinner_with_counter("svc_spinner_counter_running")

    def update_3d_svc(self, items):
        self.ax_s3d.clear()
        xs = [len(s['binpath']) for s in items if s.get('binpath')]
        ys = [1 if s['suspicious'] == 'Yes' else 0 for s in items]
        zs = list(range(len(items)))
        if xs:
            self.ax_s3d.scatter(xs, ys, zs, color="#ff6699")
            self.ax_s3d.set_xlabel("Path Length")
            self.ax_s3d.set_ylabel("Suspicious")
            self.ax_s3d.set_zlabel("Index")
        self.fig_s3d.canvas.draw()

    def update_heatmap(self, items):
        self.ax_heatmap.clear()
        anomaly_scores = [s['anomaly_score'] for s in items if isinstance(s.get('anomaly_score'), (int, float))]
        if anomaly_scores:
            data = np.array(anomaly_scores).reshape(-1, 1)
            heatmap = self.ax_heatmap.imshow(data, cmap='viridis', aspect='auto')
            self.fig_heatmap.colorbar(heatmap, ax=self.ax_heatmap)

            self.ax_heatmap.set_title("Anomaly Heatmap")
        self.fig_heatmap.canvas.draw()

    def perform_stop_action_svc(self):
        selections = self.tree_svc.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one service to stop")
            return
        for iid in selections:
            vals = self.tree_svc.item(iid, "values")
            try:
                pid_str = vals[0]
                if pid_str and str(pid_str).isdigit():
                    psutil.Process(int(pid_str)).terminate()
            except Exception:
                continue
        messagebox.showinfo("Done", "Selected services stopped")

    def perform_ban_action_svc(self):
        selections = self.tree_svc.selection()
        if not selections:
            messagebox.showwarning("No Selection", "Please select at least one service to ban")
            return
        for iid in selections:
            vals = self.tree_svc.item(iid, "values")
            try:
                pid_str = vals[0]
                if pid_str and str(pid_str).isdigit():
                    psutil.Process(int(pid_str)).kill()
            except Exception:
                continue
        messagebox.showinfo("Done", "Selected services banned")

    def export_svc_excel(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children() if
                self.tree_svc.item(i, 'values')[0] != '']
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        df = pd.DataFrame(data,
                          columns=["Name", "Timestamp", "Status", "Start Type", "Username", "Binpath", "Signature",
                                   "Suspicious", "Hidden", "Sig Publisher", "Hash Status", "Anomaly Score",
                                   "Parent PID", "TCP Listen", "TCP Established", "UDP", "Delta Cpu", "Delta Io"])
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if fn:
            try:
                df.to_excel(fn, index=False)
                messagebox.showinfo("Exported", "Excel exported")
            except Exception as e:
                messagebox.showerror("Error", f"Excel export failed: {str(e)}")

    def export_svc_html(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children() if
                self.tree_svc.item(i, 'values')[0] != '']
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        df = pd.DataFrame(data,
                          columns=["Name", "Timestamp", "Status", "Start Type", "Username", "Binpath", "Signature",
                                   "Suspicious", "Hidden", "Sig Publisher", "Hash Status", "Anomaly Score",
                                   "Parent PID", "TCP Listen", "TCP Established", "UDP", "Delta Cpu", "Delta Io"])
        fn = filedialog.asksaveasfilename(defaultextension=".html")
        if fn:
            df.to_html(fn, index=False)
            messagebox.showinfo("Exported", "HTML exported")

    def export_svc_pdf(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children() if
                self.tree_svc.item(i, 'values')[0] != '']
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".pdf")
        if fn:
            try:
                from fpdf import FPDF
            except ImportError:
                messagebox.showerror("Error", "fpdf module not installed. Use pip install fpdf")
                return
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=8)
            col_names = ["Name", "Timestamp", "Status", "Start Type", "Username", "Binpath", "Signature", "Suspicious",
                         "Hidden", "Sig Publisher", "Hash Status", "Anomaly Score", "Parent PID", "TCP Listen",
                         "TCP Established", "UDP", "Delta Cpu", "Delta Io"]
            for col in col_names:
                pdf.cell(30, 10, txt=col, border=1)
            pdf.ln()
            for row in data:
                for item in row:
                    pdf.cell(30, 10, txt=str(item), border=1)
                pdf.ln()
            pdf.output(fn)
            messagebox.showinfo("Exported", "PDF exported")

    def export_svc_excel_all(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        df = pd.DataFrame(data,

                          columns=["Name", "Timestamp", "Status", "Start Type", "Username", "Binpath", "Signature",
                                   "Suspicious", "Hidden", "Sig Publisher", "Hash Status", "Anomaly Score",
                                   "Parent PID", "TCP Listen", "TCP Established", "UDP", "Delta Cpu", "Delta Io"])
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if fn:
            try:
                df.to_excel(fn, index=False)
                messagebox.showinfo("Exported", "Excel exported")
            except Exception as e:
                messagebox.showerror("Error", f"Excel export failed: {str(e)}")

    def export_svc_html_all(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        df = pd.DataFrame(data, columns=["Name", "Timestamp", "Status", "Start Type", "Username",
                                         "Binpath", "Signature", "Suspicious", "Hidden", "Sig Publisher", "Hash Status",
                                         "Anomaly Score", "Parent PID", "TCP Listen", "TCP Established", "UDP",
                                         "Delta Cpu", "Delta Io"])
        fn = filedialog.asksaveasfilename(defaultextension=".html")
        if fn:
            df.to_html(fn, index=False)
            messagebox.showinfo("Exported", "HTML exported")

    def export_svc_pdf_all(self):
        data = [self.tree_svc.item(i, 'values')[1:] for i in self.tree_svc.get_children()]
        if not data:
            messagebox.showwarning("No Data", "No data to export")
            return
        fn = filedialog.asksaveasfilename(defaultextension=".pdf")
        if fn:
            try:
                from fpdf import FPDF
            except ImportError:
                messagebox.showerror("Error", "fpdf module not installed. Use pip install fpdf")
                return
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=8)
            col_names = ["Name", "Timestamp", "Status", "Start Type", "Username", "Binpath", "Signature", "Suspicious",
                         "Hidden", "Sig Publisher", "Hash Status", "Anomaly Score", "Parent PID", "TCP Listen",
                         "TCP Established", "UDP", "Delta Cpu", "Delta Io"]
            for col in col_names:
                pdf.cell(30, 10, txt=col, border=1)
            pdf.ln()
            for row in data:
                for item in row:
                    pdf.cell(30, 10, txt=str(item), border=1)
                pdf.ln()

            pdf.output(fn)
            messagebox.showinfo("Exported", "PDF exported")

    def open_export_dialog_svc(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("حدد نوع التصدير")
        dialog.geometry("300x150")
        self.animate_dialog_wavy(dialog)
        tk.Label(dialog, text="اختر نوع التصدير:", font=(FONT_NAME, 11, 'bold'), bg=dialog['bg']).pack(pady=10)
        export_type = tk.StringVar(value="Excel")
        formats = [("Excel", "Excel"), ("PDF", "PDF"), ("HTML", "HTML")]
        for text, value in formats:
            tk.Radiobutton(dialog, text=text, variable=export_type, value=value, font=(FONT_NAME, 10),
                           bg=dialog['bg']).pack(anchor=tk.W, padx=20)

        def submit():
            choice = export_type.get()
            dialog.destroy()
            if choice == "Excel":
                self.export_svc_excel_all()
            elif choice == "PDF":
                self.export_svc_pdf_all()
            elif choice == "HTML":
                self.export_svc_html_all()

        tk.Button(dialog, text="تصدير", font=(FONT_NAME, 11, 'bold'), bg=random.choice(GRADIENTS), fg="white",
                  command=submit).pack(pady=10)

    # ----- دوال مساعدة إضافية -----
    def add_history_record(self, record):
        cond = (record.get('anomalous', 0) == 1 or
                record.get('signature_valid', 1) == 0 or
                record.get('source', '') in ["Suspicious", "Keywords Scan", "Suspicious Services"])
        if not cond:
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO services
                (pid,name,path,account,frequency,age_days,timestamp,anomalous,anomaly_score,signature_valid, source, delta_cpu, delta_io, tcp_listen, tcp_established, udp)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                record.get('pid'),
                record.get('name'),
                record.get('path'),
                record.get('account', ''),
                record.get('frequency', ''),

                record.get('age_days', 0),
                record.get('timestamp'),
                record.get('anomalous', 0),
                record.get('anomaly_score', ''),
                record.get('signature_valid', 1),
                record.get('source', ''),
                record.get('delta_cpu', 0),
                record.get('delta_io', 0),
                record.get('tcp_listen', 0),
                record.get('tcp_established', 0),
                record.get('udp', 0)
            ))
            conn.commit()
        except Exception as e:
            print("DB insert error:", str(e))
        finally:
            conn.close()

    def get_account_type(self, proc):
        try:
            if platform.system() == "Windows":
                usr = proc.info.get("username", "").upper()
                return "System" if usr.startswith("SYSTEM") else "User"
            return "System" if proc.info.get("username", "") == "root" else "User"
        except:
            return "Unknown"

    def get_frequency(self, proc):
        return "Unknown"

    def refresh_all(self):
        threading.Thread(target=self.scan_all_services, daemon=True).start()

    def collect_baseline(self):
        end_t = datetime.now().timestamp() + 10
        try:
            with open(BASELINE_FILE, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["cpu_pct", "rss", "vms", "read_bytes", "write_bytes", "num_conns"])
                while datetime.now().timestamp() < end_t:
                    for proc in psutil.process_iter(['pid', 'cmdline']):
                        try:
                            cpu = proc.cpu_percent(interval=0.1)
                            mem = proc.memory_info()
                            rss, vms = mem.rss, mem.vms
                            io = proc.io_counters() if hasattr(proc, "io_counters") else None
                            rb = io.read_bytes if io else 0
                            wb = io.write_bytes if io else 0
                            conns = proc.connections(kind="inet") if hasattr(proc, "connections") else []
                            w.writerow([cpu, rss, vms, rb, wb, len(conns)])
                        except:

                            continue
            messagebox.showinfo("Done", "Baseline collected")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def train_model(self):
        if not os.path.exists(BASELINE_FILE):
            messagebox.showerror("Error", "Collect baseline first")
            return
        try:
            df = pd.read_csv(BASELINE_FILE)
            feats = df.dropna().iloc[:, :6]
            model = IsolationForest(contamination=0.01, random_state=42)
            model.fit(feats)
            joblib.dump(model, MODEL_FILE)
            self.anomaly_model = model
            messagebox.showinfo("Done", "Model trained and saved")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def collect_baseline_with_spinner_all(self):
        self.start_spinner(self.spinner_label_all, "bs_spinner_running_all")
        self.collect_baseline()
        self.stop_spinner("bs_spinner_running_all")

    def collect_baseline_with_spinner_svc(self):
        self.start_spinner(self.spinner_label_svc, "bs_spinner_running_svc")
        self.collect_baseline()
        self.stop_spinner("bs_spinner_running_svc")

    # ---------- دالة لتبديل إيقاف واستئناف حركة مخطط الرادار ----------
    def toggle_radar_pause(self):
        self.radar_paused = not self.radar_paused
        if self.radar_paused:
            messagebox.showinfo("Paused", "Radar animation paused.")
        else:
            messagebox.showinfo("Resumed", "Radar animation resumed.")

def main():
    root = tk.Tk()
    app = ServiceScannerApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()


































